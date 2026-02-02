"""
Excel Hyperlink Extractor Module
================================
Extracts hyperlinks from Excel files (.xlsx, .xls) for validation.

This module provides functionality to extract all hyperlinks from Excel
workbooks, including:
- Cell hyperlinks (explicit hyperlink objects)
- HYPERLINK() formula-based links
- Email addresses and URLs in cell values

Requires: openpyxl for .xlsx files
Optional: xlrd for legacy .xls files (not included by default)
"""

import re
import os
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple, Any
from enum import Enum

# Try to import openpyxl for .xlsx files
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Try to import xlrd for legacy .xls files
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


class LinkSource(Enum):
    """Source type for extracted links."""
    HYPERLINK_OBJECT = "hyperlink"      # Cell has explicit hyperlink
    HYPERLINK_FORMULA = "formula"        # HYPERLINK() function
    CELL_VALUE = "cell_value"            # URL/email in cell text
    COMMENT = "comment"                  # Link in cell comment


@dataclass
class ExtractedExcelLink:
    """Represents a hyperlink extracted from an Excel file."""
    url: str
    display_text: str
    sheet_name: str
    cell_address: str
    row: int
    column: int
    source: LinkSource
    tooltip: Optional[str] = None
    context: Optional[str] = None  # Surrounding cell values for context

    def to_dict(self) -> Dict[str, Any]:
        """Convert to dictionary for JSON serialization."""
        return {
            'url': self.url,
            'display_text': self.display_text,
            'sheet_name': self.sheet_name,
            'cell_address': self.cell_address,
            'row': self.row,
            'column': self.column,
            'source': self.source.value,
            'tooltip': self.tooltip,
            'context': self.context
        }


@dataclass
class SheetSummary:
    """Summary of links found in a single worksheet."""
    name: str
    total_links: int = 0
    hyperlink_objects: int = 0
    formula_links: int = 0
    cell_value_links: int = 0

    def to_dict(self) -> Dict[str, Any]:
        return {
            'name': self.name,
            'total_links': self.total_links,
            'hyperlink_objects': self.hyperlink_objects,
            'formula_links': self.formula_links,
            'cell_value_links': self.cell_value_links
        }


@dataclass
class ExcelExtractionResult:
    """Complete result of Excel hyperlink extraction."""
    file_path: str
    file_name: str
    total_links: int = 0
    sheets_processed: int = 0
    links: List[ExtractedExcelLink] = field(default_factory=list)
    sheet_summaries: List[SheetSummary] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            'file_path': self.file_path,
            'file_name': self.file_name,
            'total_links': self.total_links,
            'sheets_processed': self.sheets_processed,
            'links': [link.to_dict() for link in self.links],
            'sheet_summaries': [s.to_dict() for s in self.sheet_summaries],
            'errors': self.errors
        }

    def get_urls(self) -> List[str]:
        """Get just the URLs as a list."""
        return [link.url for link in self.links]

    def get_unique_urls(self) -> List[str]:
        """Get unique URLs."""
        seen = set()
        unique = []
        for link in self.links:
            if link.url not in seen:
                seen.add(link.url)
                unique.append(link.url)
        return unique


class ExcelExtractor:
    """
    Extracts hyperlinks from Excel files.

    Supports:
    - .xlsx files (requires openpyxl)
    - .xls files (requires xlrd, optional)
    """

    # Regex patterns for detecting URLs and emails in cell values
    URL_PATTERN = re.compile(
        r'https?://[^\s<>"\']+|'
        r'www\.[^\s<>"\']+|'
        r'ftp://[^\s<>"\']+'
    )

    EMAIL_PATTERN = re.compile(
        r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}'
    )

    # HYPERLINK formula pattern
    HYPERLINK_FORMULA_PATTERN = re.compile(
        r'=HYPERLINK\s*\(\s*"([^"]+)"(?:\s*,\s*"([^"]*)")?\s*\)',
        re.IGNORECASE
    )

    def __init__(self, extract_from_values: bool = True,
                 extract_from_formulas: bool = True,
                 extract_from_comments: bool = False):
        """
        Initialize the Excel extractor.

        Args:
            extract_from_values: Also scan cell values for URLs/emails
            extract_from_formulas: Extract links from HYPERLINK() formulas
            extract_from_comments: Also scan cell comments for URLs
        """
        self.extract_from_values = extract_from_values
        self.extract_from_formulas = extract_from_formulas
        self.extract_from_comments = extract_from_comments

    def extract(self, file_path: str) -> ExcelExtractionResult:
        """
        Extract all hyperlinks from an Excel file.

        Args:
            file_path: Path to the Excel file

        Returns:
            ExcelExtractionResult with all extracted links
        """
        result = ExcelExtractionResult(
            file_path=file_path,
            file_name=os.path.basename(file_path)
        )

        # Determine file type
        ext = os.path.splitext(file_path)[1].lower()

        if ext == '.xlsx':
            if not OPENPYXL_AVAILABLE:
                result.errors.append("openpyxl library not installed. Cannot process .xlsx files.")
                return result
            self._extract_xlsx(file_path, result)

        elif ext == '.xls':
            if not XLRD_AVAILABLE:
                result.errors.append("xlrd library not installed. Cannot process legacy .xls files.")
                return result
            self._extract_xls(file_path, result)

        else:
            result.errors.append(f"Unsupported file format: {ext}")

        return result

    def _extract_xlsx(self, file_path: str, result: ExcelExtractionResult) -> None:
        """Extract hyperlinks from .xlsx file using openpyxl."""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=False)

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_summary = SheetSummary(name=sheet_name)

                # Process hyperlink objects
                if hasattr(ws, 'hyperlinks'):
                    for hyperlink in ws.hyperlinks:
                        if hyperlink.target:
                            cell = ws[hyperlink.ref]
                            display_text = str(cell.value) if cell.value else hyperlink.target

                            link = ExtractedExcelLink(
                                url=hyperlink.target,
                                display_text=display_text,
                                sheet_name=sheet_name,
                                cell_address=hyperlink.ref,
                                row=cell.row,
                                column=cell.column,
                                source=LinkSource.HYPERLINK_OBJECT,
                                tooltip=hyperlink.tooltip,
                                context=self._get_cell_context(ws, cell.row, cell.column)
                            )
                            result.links.append(link)
                            sheet_summary.hyperlink_objects += 1

                # Process cells for formulas and values
                for row in ws.iter_rows():
                    for cell in row:
                        # Check for HYPERLINK formula
                        if self.extract_from_formulas and cell.value:
                            cell_value = str(cell.value)
                            if cell_value.startswith('=HYPERLINK'):
                                match = self.HYPERLINK_FORMULA_PATTERN.match(cell_value)
                                if match:
                                    url = match.group(1)
                                    display = match.group(2) if match.group(2) else url

                                    link = ExtractedExcelLink(
                                        url=url,
                                        display_text=display,
                                        sheet_name=sheet_name,
                                        cell_address=cell.coordinate,
                                        row=cell.row,
                                        column=cell.column,
                                        source=LinkSource.HYPERLINK_FORMULA,
                                        context=self._get_cell_context(ws, cell.row, cell.column)
                                    )
                                    result.links.append(link)
                                    sheet_summary.formula_links += 1

                        # Check cell values for URLs/emails
                        if self.extract_from_values and cell.value and not str(cell.value).startswith('='):
                            cell_text = str(cell.value)

                            # Skip if this cell already has a hyperlink object
                            has_hyperlink = any(
                                link.cell_address == cell.coordinate and
                                link.sheet_name == sheet_name and
                                link.source == LinkSource.HYPERLINK_OBJECT
                                for link in result.links
                            )

                            if not has_hyperlink:
                                # Find URLs
                                for url_match in self.URL_PATTERN.finditer(cell_text):
                                    url = url_match.group()
                                    link = ExtractedExcelLink(
                                        url=url,
                                        display_text=cell_text[:100],
                                        sheet_name=sheet_name,
                                        cell_address=cell.coordinate,
                                        row=cell.row,
                                        column=cell.column,
                                        source=LinkSource.CELL_VALUE,
                                        context=self._get_cell_context(ws, cell.row, cell.column)
                                    )
                                    result.links.append(link)
                                    sheet_summary.cell_value_links += 1

                                # Find emails (only if no URL found in this cell)
                                if not self.URL_PATTERN.search(cell_text):
                                    for email_match in self.EMAIL_PATTERN.finditer(cell_text):
                                        email = email_match.group()
                                        link = ExtractedExcelLink(
                                            url=f"mailto:{email}",
                                            display_text=email,
                                            sheet_name=sheet_name,
                                            cell_address=cell.coordinate,
                                            row=cell.row,
                                            column=cell.column,
                                            source=LinkSource.CELL_VALUE,
                                            context=self._get_cell_context(ws, cell.row, cell.column)
                                        )
                                        result.links.append(link)
                                        sheet_summary.cell_value_links += 1

                        # Check comments
                        if self.extract_from_comments and cell.comment:
                            comment_text = str(cell.comment.text)
                            for url_match in self.URL_PATTERN.finditer(comment_text):
                                url = url_match.group()
                                link = ExtractedExcelLink(
                                    url=url,
                                    display_text=f"[Comment] {url}",
                                    sheet_name=sheet_name,
                                    cell_address=cell.coordinate,
                                    row=cell.row,
                                    column=cell.column,
                                    source=LinkSource.COMMENT,
                                    context=comment_text[:200]
                                )
                                result.links.append(link)

                sheet_summary.total_links = (
                    sheet_summary.hyperlink_objects +
                    sheet_summary.formula_links +
                    sheet_summary.cell_value_links
                )
                result.sheet_summaries.append(sheet_summary)
                result.sheets_processed += 1

            wb.close()
            result.total_links = len(result.links)

        except Exception as e:
            result.errors.append(f"Error reading Excel file: {str(e)}")

    def _extract_xls(self, file_path: str, result: ExcelExtractionResult) -> None:
        """Extract hyperlinks from legacy .xls file using xlrd."""
        try:
            wb = xlrd.open_workbook(file_path, formatting_info=True)

            for sheet_idx in range(wb.nsheets):
                ws = wb.sheet_by_index(sheet_idx)
                sheet_summary = SheetSummary(name=ws.name)

                # Get hyperlinks from the sheet
                if hasattr(ws, 'hyperlink_map'):
                    for (row, col), hyperlink in ws.hyperlink_map.items():
                        if hyperlink.url_or_path:
                            cell_value = ws.cell_value(row, col)
                            display_text = str(cell_value) if cell_value else hyperlink.url_or_path

                            link = ExtractedExcelLink(
                                url=hyperlink.url_or_path,
                                display_text=display_text,
                                sheet_name=ws.name,
                                cell_address=f"{get_column_letter_xlrd(col + 1)}{row + 1}",
                                row=row + 1,
                                column=col + 1,
                                source=LinkSource.HYPERLINK_OBJECT,
                                tooltip=hyperlink.desc if hasattr(hyperlink, 'desc') else None
                            )
                            result.links.append(link)
                            sheet_summary.hyperlink_objects += 1

                # Scan cell values for URLs
                if self.extract_from_values:
                    for row in range(ws.nrows):
                        for col in range(ws.ncols):
                            cell_value = ws.cell_value(row, col)
                            if cell_value and isinstance(cell_value, str):
                                # Skip if already has hyperlink
                                has_hyperlink = hasattr(ws, 'hyperlink_map') and (row, col) in ws.hyperlink_map

                                if not has_hyperlink:
                                    # Find URLs
                                    for url_match in self.URL_PATTERN.finditer(cell_value):
                                        url = url_match.group()
                                        link = ExtractedExcelLink(
                                            url=url,
                                            display_text=cell_value[:100],
                                            sheet_name=ws.name,
                                            cell_address=f"{get_column_letter_xlrd(col + 1)}{row + 1}",
                                            row=row + 1,
                                            column=col + 1,
                                            source=LinkSource.CELL_VALUE
                                        )
                                        result.links.append(link)
                                        sheet_summary.cell_value_links += 1

                                    # Find emails
                                    if not self.URL_PATTERN.search(cell_value):
                                        for email_match in self.EMAIL_PATTERN.finditer(cell_value):
                                            email = email_match.group()
                                            link = ExtractedExcelLink(
                                                url=f"mailto:{email}",
                                                display_text=email,
                                                sheet_name=ws.name,
                                                cell_address=f"{get_column_letter_xlrd(col + 1)}{row + 1}",
                                                row=row + 1,
                                                column=col + 1,
                                                source=LinkSource.CELL_VALUE
                                            )
                                            result.links.append(link)
                                            sheet_summary.cell_value_links += 1

                sheet_summary.total_links = (
                    sheet_summary.hyperlink_objects +
                    sheet_summary.cell_value_links
                )
                result.sheet_summaries.append(sheet_summary)
                result.sheets_processed += 1

            result.total_links = len(result.links)

        except Exception as e:
            result.errors.append(f"Error reading legacy Excel file: {str(e)}")

    def _get_cell_context(self, ws, row: int, col: int, context_size: int = 1) -> str:
        """Get surrounding cell values for context."""
        context_parts = []

        # Get values from adjacent cells (left and right)
        for c_offset in range(-context_size, context_size + 1):
            if c_offset == 0:
                continue
            try:
                adj_col = col + c_offset
                if adj_col >= 1:
                    cell = ws.cell(row=row, column=adj_col)
                    if cell.value:
                        context_parts.append(str(cell.value)[:50])
            except:
                pass

        return " | ".join(context_parts) if context_parts else None


def get_column_letter_xlrd(col_num: int) -> str:
    """Convert column number to letter (for xlrd compatibility)."""
    result = ""
    while col_num > 0:
        col_num, remainder = divmod(col_num - 1, 26)
        result = chr(65 + remainder) + result
    return result


# Convenience functions
def extract_excel_links(file_path: str,
                        extract_from_values: bool = True,
                        extract_from_formulas: bool = True) -> ExcelExtractionResult:
    """
    Extract all hyperlinks from an Excel file.

    Args:
        file_path: Path to the Excel file (.xlsx or .xls)
        extract_from_values: Also scan cell values for URLs/emails
        extract_from_formulas: Extract links from HYPERLINK() formulas

    Returns:
        ExcelExtractionResult containing all extracted links
    """
    extractor = ExcelExtractor(
        extract_from_values=extract_from_values,
        extract_from_formulas=extract_from_formulas
    )
    return extractor.extract(file_path)


def get_urls_from_excel(file_path: str) -> List[str]:
    """
    Get just the URLs from an Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        List of URLs found in the file
    """
    result = extract_excel_links(file_path)
    return result.get_urls()


def is_excel_available() -> Dict[str, bool]:
    """Check which Excel formats are supported."""
    return {
        'xlsx': OPENPYXL_AVAILABLE,
        'xls': XLRD_AVAILABLE
    }
