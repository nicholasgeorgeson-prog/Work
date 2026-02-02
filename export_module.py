#!/usr/bin/env python3
"""
TechWriterReview Export Module
===============================
Export functionality for:
- Excel reports with formatting
- CSV exports
- PDF summary reports
- JSON exports
- Compliance matrices

Created by Nicholas Georgeson
"""

import csv
import json
import io
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
from collections import defaultdict

# Excel export using openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.label import DataLabelList
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# PDF export using reportlab
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.platypus import Image as RLImage
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False


class ExcelExporter:
    """Export analysis results to Excel format."""
    
    # Color scheme
    COLORS = {
        'header_bg': 'D5E8F0',
        'header_font': '1E3A5F',
        'critical': 'DC3545',
        'high': 'FD7E14',
        'medium': 'FFC107',
        'low': '28A745',
        'info': '17A2B8',
        'alt_row': 'F8F9FA'
    }
    
    SEVERITY_ORDER = ['Critical', 'High', 'Medium', 'Low', 'Info']
    
    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
    
    def export(self, results: Dict, filename: str = None, 
               include_charts: bool = True,
               include_roles: bool = True,
               include_readability: bool = True,
               severities: List[str] = None,
               document_metadata: Dict = None) -> bytes:
        """Export full analysis results to Excel.
        
        v3.0.33 Chunk D: Enhanced export with:
        - Action Item column for reviewers
        - Timestamp in filename
        - Document metadata header
        - Severity filter support
        
        Args:
            results: Analysis results dictionary
            filename: Optional output filename
            include_charts: Include charts sheet
            include_roles: Include roles sheet
            include_readability: Include readability sheet
            severities: Optional list of severities to include (e.g., ['Critical', 'High'])
            document_metadata: Optional dict with document name, scan_date, score
        """
        # Store metadata for use in sheets
        self._document_metadata = document_metadata or {}
        self._export_timestamp = datetime.now()
        
        # Filter issues by severity if specified
        filtered_results = results.copy()
        if severities and 'issues' in filtered_results:
            filtered_results['issues'] = [
                issue for issue in filtered_results.get('issues', [])
                if issue.get('severity', 'Info') in severities
            ]
        
        # Summary sheet
        self._create_summary_sheet(filtered_results)
        
        # Issues sheet
        self._create_issues_sheet(filtered_results.get('issues', []))
        
        # By Category breakdown
        self._create_category_sheet(filtered_results.get('issues', []))
        
        # Roles sheet (if available)
        if include_roles and filtered_results.get('roles'):
            self._create_roles_sheet(filtered_results['roles'])
        
        # Readability sheet
        if include_readability and filtered_results.get('readability'):
            self._create_readability_sheet(filtered_results['readability'])
        
        # Charts sheet
        if include_charts:
            self._create_charts_sheet(filtered_results)
        
        # Save to bytes
        output = io.BytesIO()
        self.wb.save(output)
        output.seek(0)
        
        if filename:
            with open(filename, 'wb') as f:
                f.write(output.getvalue())
        
        return output.getvalue()
    
    def _create_summary_sheet(self, results: Dict):
        """Create summary sheet with document metadata header.
        
        v3.0.33 Chunk D: Added metadata header row showing:
        - Document filename
        - Scan timestamp  
        - Quality score
        """
        ws = self.wb.create_sheet("Summary")
        
        # Header styling
        header_font = Font(bold=True, color=self.COLORS['header_font'], size=14)
        header_fill = PatternFill(start_color=self.COLORS['header_bg'], 
                                  end_color=self.COLORS['header_bg'], fill_type='solid')
        
        # Title
        ws['A1'] = 'TechWriterReview Analysis Report'
        ws['A1'].font = Font(bold=True, size=18)
        ws.merge_cells('A1:D1')
        
        # v3.0.33 Chunk D: Document metadata header
        meta = self._document_metadata if hasattr(self, '_document_metadata') else {}
        export_ts = self._export_timestamp if hasattr(self, '_export_timestamp') else datetime.now()
        
        ws['A2'] = f"Document: {meta.get('filename', results.get('document_info', {}).get('filename', 'Unknown'))}"
        ws['A2'].font = Font(italic=True, color='666666')
        
        ws['A3'] = f"Scan Date: {meta.get('scan_date', export_ts.strftime('%Y-%m-%d %H:%M:%S'))}"
        ws['A3'].font = Font(italic=True, color='666666')
        
        ws['A4'] = f"Quality Score: {meta.get('score', results.get('score', 100))}"
        ws['A4'].font = Font(italic=True, color='666666')
        
        ws['A5'] = f"Export Generated: {export_ts.strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A5'].font = Font(italic=True, color='666666')
        
        # Document info section (shifted down)
        row = 7
        ws[f'A{row}'] = 'Document Information'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        doc_info = results.get('document_info', {})
        info_items = [
            ('Filename', doc_info.get('filename', 'N/A')),
            ('Word Count', doc_info.get('word_count', 'N/A')),
            ('Paragraphs', doc_info.get('paragraph_count', 'N/A')),
            ('Tables', doc_info.get('table_count', 0)),
            ('Figures', doc_info.get('figure_count', 0)),
        ]
        
        row += 1
        for label, value in info_items:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Score section
        row += 1
        ws[f'A{row}'] = 'Quality Score'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        ws[f'A{row}'] = 'Overall Score'
        ws[f'B{row}'] = results.get('score', 100)
        ws[f'B{row}'].font = Font(bold=True, size=16)
        
        row += 1
        ws[f'A{row}'] = 'Grade'
        ws[f'B{row}'] = results.get('grade', 'A')
        ws[f'B{row}'].font = Font(bold=True, size=16)
        
        row += 1
        ws[f'A{row}'] = 'Total Issues'
        ws[f'B{row}'] = results.get('issue_count', 0)
        
        # Severity breakdown
        row += 2
        ws[f'A{row}'] = 'Issues by Severity'
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        by_severity = results.get('by_severity', {})
        row += 1
        for sev in self.SEVERITY_ORDER:
            ws[f'A{row}'] = sev
            ws[f'B{row}'] = by_severity.get(sev, 0)
            
            # Color the severity
            color = self.COLORS.get(sev.lower(), 'FFFFFF')
            ws[f'A{row}'].fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            if sev in ['Critical', 'High']:
                ws[f'A{row}'].font = Font(color='FFFFFF')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
    
    def _create_issues_sheet(self, issues: List[Dict]):
        """Create detailed issues sheet with provenance data and Action Item column.
        
        v3.0.33 Chunk D: Added editable "Action Item" column for reviewer notes.
        """
        ws = self.wb.create_sheet("Issues")
        
        # Headers - now includes provenance columns and Action Item
        # v3.0.33 Chunk D: Added 'Action Item' as last column
        headers = ['#', 'Severity', 'Category', 'Message', 'Flagged Text', 'Suggestion', 'Location', 'Validated', 'Original Text', 'Action Item']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=self.COLORS['header_font'], 
                                  end_color=self.COLORS['header_font'], fill_type='solid')
        
        # v3.0.33 Chunk D: Action Item column styling (yellow highlight)
        action_header_fill = PatternFill(start_color='FFC107', end_color='FFC107', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            if header == 'Action Item':
                cell.fill = action_header_fill
                cell.font = Font(bold=True, color='000000')
            else:
                cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Sort issues by severity
        severity_order = {s: i for i, s in enumerate(self.SEVERITY_ORDER)}
        sorted_issues = sorted(issues, key=lambda x: severity_order.get(x.get('severity', 'Info'), 99))
        
        # Add issues
        for row_num, issue in enumerate(sorted_issues, 2):
            severity = issue.get('severity', 'Info')
            
            ws.cell(row=row_num, column=1, value=row_num - 1)
            
            sev_cell = ws.cell(row=row_num, column=2, value=severity)
            color = self.COLORS.get(severity.lower(), 'FFFFFF')
            sev_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            if severity in ['Critical', 'High']:
                sev_cell.font = Font(color='FFFFFF')
            
            ws.cell(row=row_num, column=3, value=issue.get('category', ''))
            ws.cell(row=row_num, column=4, value=issue.get('message', ''))
            ws.cell(row=row_num, column=5, value=issue.get('flagged_text', issue.get('text', ''))[:200])
            ws.cell(row=row_num, column=6, value=issue.get('suggestion', ''))
            ws.cell(row=row_num, column=7, value=f"Para {issue.get('paragraph_index', 'N/A')}")
            
            # Provenance columns
            source = issue.get('source', {})
            is_validated = source.get('is_validated', False) if source else False
            original_text = source.get('original_text', '') if source else ''
            
            validated_cell = ws.cell(row=row_num, column=8, value='Yes' if is_validated else 'No')
            if is_validated:
                validated_cell.fill = PatternFill(start_color='E9D5FF', end_color='E9D5FF', fill_type='solid')  # Light purple
            
            ws.cell(row=row_num, column=9, value=original_text[:100] if original_text else '')
            
            # v3.0.33 Chunk D: Action Item column - empty, editable cell with light yellow background
            action_cell = ws.cell(row=row_num, column=10, value='')
            action_cell.fill = PatternFill(start_color='FFF9E6', end_color='FFF9E6', fill_type='solid')
            action_cell.alignment = Alignment(horizontal='left', wrap_text=True)
            
            # Alternate row coloring
            if row_num % 2 == 0:
                for col in range(1, 10):  # Don't override Action Item column color
                    if col not in [2, 8]:  # Don't override severity or validated color
                        current_cell = ws.cell(row=row_num, column=col)
                        if not current_cell.fill or current_cell.fill.fgColor.rgb == '00000000':
                            current_cell.fill = PatternFill(
                                start_color=self.COLORS['alt_row'],
                                end_color=self.COLORS['alt_row'],
                                fill_type='solid'
                            )
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 50
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 40
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10
        ws.column_dimensions['I'].width = 30
        ws.column_dimensions['J'].width = 30  # v3.0.33 Chunk D: Action Item column
        
        # Add filters (expanded to include Action Item)
        ws.auto_filter.ref = f"A1:J{len(issues) + 1}"
    
    def _create_category_sheet(self, issues: List[Dict]):
        """Create category breakdown sheet."""
        ws = self.wb.create_sheet("By Category")
        
        # Count by category
        by_category = defaultdict(lambda: {'total': 0, 'by_severity': defaultdict(int)})
        for issue in issues:
            cat = issue.get('category', 'Other')
            sev = issue.get('severity', 'Info')
            by_category[cat]['total'] += 1
            by_category[cat]['by_severity'][sev] += 1
        
        # Headers
        headers = ['Category', 'Total'] + self.SEVERITY_ORDER
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color=self.COLORS['header_font'], 
                                  end_color=self.COLORS['header_font'], fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        sorted_cats = sorted(by_category.items(), key=lambda x: -x[1]['total'])
        for row_num, (cat, data) in enumerate(sorted_cats, 2):
            ws.cell(row=row_num, column=1, value=cat)
            ws.cell(row=row_num, column=2, value=data['total'])
            for col, sev in enumerate(self.SEVERITY_ORDER, 3):
                ws.cell(row=row_num, column=col, value=data['by_severity'].get(sev, 0))
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 10
    
    def _create_roles_sheet(self, roles: Dict):
        """Create roles breakdown sheet."""
        ws = self.wb.create_sheet("Roles")
        
        # Headers
        headers = ['Role Name', 'Occurrences', 'Responsibilities', 'Action Types']
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Data rows
        row = 2
        for role_name, role_data in roles.items():
            if isinstance(role_data, dict):
                ws.cell(row=row, column=1, value=role_name)
                ws.cell(row=row, column=2, value=role_data.get('count', 1))
                
                # Responsibilities
                resps = role_data.get('responsibilities', [])
                if resps:
                    ws.cell(row=row, column=3, value='; '.join(str(r) for r in resps[:5]))
                
                # Action types
                actions = role_data.get('action_types', {})
                if actions:
                    action_str = ', '.join(f"{k}: {v}" for k, v in list(actions.items())[:5])
                    ws.cell(row=row, column=4, value=action_str)
                
                row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 50
        ws.column_dimensions['D'].width = 40
    
    def _create_readability_sheet(self, readability: Dict):
        """Create readability metrics sheet."""
        ws = self.wb.create_sheet("Readability")
        
        # Title
        ws['A1'] = 'Readability Metrics'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        # Metrics
        metrics = [
            ('Flesch Reading Ease', readability.get('flesch_reading_ease'), 
             'Higher = easier to read (0-100)'),
            ('Flesch-Kincaid Grade', readability.get('flesch_kincaid_grade'),
             'US school grade level'),
            ('Gunning Fog Index', readability.get('gunning_fog'),
             'Years of education needed'),
            ('SMOG Index', readability.get('smog_index'),
             'Years of education needed'),
            ('Coleman-Liau Index', readability.get('coleman_liau'),
             'US school grade level'),
            ('Automated Readability Index', readability.get('automated_readability'),
             'US school grade level'),
        ]
        
        row = 3
        for metric, value, description in metrics:
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=round(value, 1) if value else 'N/A')
            ws.cell(row=row, column=3, value=description)
            ws.cell(row=row, column=3).font = Font(italic=True, color='666666')
            row += 1
        
        # Interpretation guide
        row += 2
        ws.cell(row=row, column=1, value='Flesch Reading Ease Interpretation')
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        interpretations = [
            ('90-100', 'Very Easy', '5th grade'),
            ('80-90', 'Easy', '6th grade'),
            ('70-80', 'Fairly Easy', '7th grade'),
            ('60-70', 'Standard', '8th-9th grade'),
            ('50-60', 'Fairly Difficult', '10th-12th grade'),
            ('30-50', 'Difficult', 'College'),
            ('0-30', 'Very Difficult', 'College Graduate'),
        ]
        
        row += 1
        for score_range, difficulty, grade in interpretations:
            ws.cell(row=row, column=1, value=score_range)
            ws.cell(row=row, column=2, value=difficulty)
            ws.cell(row=row, column=3, value=grade)
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
    
    def _create_charts_sheet(self, results: Dict):
        """Create charts sheet (placeholder - actual charts would need more implementation)."""
        ws = self.wb.create_sheet("Charts")
        
        ws['A1'] = 'Charts'
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = 'Note: Visual charts are displayed in the web interface.'
        ws['A4'] = 'This sheet contains the raw data for chart generation.'
        
        # Severity data for charts
        ws['A6'] = 'Severity Distribution'
        ws['A6'].font = Font(bold=True)
        
        by_severity = results.get('by_severity', {})
        row = 7
        for sev in self.SEVERITY_ORDER:
            ws.cell(row=row, column=1, value=sev)
            ws.cell(row=row, column=2, value=by_severity.get(sev, 0))
            row += 1
        
        # Category data
        row += 2
        ws.cell(row=row, column=1, value='Category Distribution')
        ws.cell(row=row, column=1).font = Font(bold=True)
        
        by_category = results.get('by_category', {})
        row += 1
        for cat, count in sorted(by_category.items(), key=lambda x: -x[1])[:15]:
            ws.cell(row=row, column=1, value=cat)
            ws.cell(row=row, column=2, value=count)
            row += 1


class CSVExporter:
    """Export analysis results to CSV format."""
    
    @staticmethod
    def export_issues(issues: List[Dict], filename: str = None) -> str:
        """Export issues to CSV."""
        output = io.StringIO()
        
        fieldnames = ['#', 'Severity', 'Category', 'Message', 'Flagged Text', 
                     'Suggestion', 'Paragraph', 'Start', 'End']
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for i, issue in enumerate(issues, 1):
            writer.writerow({
                '#': i,
                'Severity': issue.get('severity', ''),
                'Category': issue.get('category', ''),
                'Message': issue.get('message', ''),
                'Flagged Text': issue.get('flagged_text', issue.get('text', '')),
                'Suggestion': issue.get('suggestion', ''),
                'Paragraph': issue.get('paragraph_index', ''),
                'Start': issue.get('start_offset', ''),
                'End': issue.get('end_offset', '')
            })
        
        csv_content = output.getvalue()
        
        if filename:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                f.write(csv_content)
        
        return csv_content
    
    @staticmethod
    def export_roles(roles: Dict, filename: str = None) -> str:
        """Export roles to CSV."""
        output = io.StringIO()
        
        fieldnames = ['Role Name', 'Occurrences', 'Responsibilities', 'Action Types', 'Variants']
        
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        
        for role_name, role_data in roles.items():
            if isinstance(role_data, dict):
                writer.writerow({
                    'Role Name': role_name,
                    'Occurrences': role_data.get('count', 1),
                    'Responsibilities': '; '.join(str(r) for r in role_data.get('responsibilities', [])[:10]),
                    'Action Types': ', '.join(f"{k}:{v}" for k, v in role_data.get('action_types', {}).items()),
                    'Variants': ', '.join(role_data.get('variants', []))
                })
        
        csv_content = output.getvalue()
        
        if filename:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                f.write(csv_content)
        
        return csv_content


class PDFExporter:
    """Export analysis results to PDF format."""
    
    def __init__(self):
        if not PDF_AVAILABLE:
            raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
        
        self.styles = getSampleStyleSheet()
        self._setup_styles()
    
    def _setup_styles(self):
        """Setup custom styles."""
        self.styles.add(ParagraphStyle(
            name='Title',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30
        ))
        self.styles.add(ParagraphStyle(
            name='SectionHeader',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            spaceBefore=20,
            textColor=colors.HexColor('#1E3A5F')
        ))
    
    def export(self, results: Dict, filename: str) -> bytes:
        """Export analysis to PDF."""
        output = io.BytesIO()
        doc = SimpleDocTemplate(
            output,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch
        )
        
        story = []
        
        # Title
        story.append(Paragraph('TechWriterReview Analysis Report', self.styles['Title']))
        story.append(Paragraph(
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            self.styles['Normal']
        ))
        story.append(Spacer(1, 20))
        
        # Summary section
        story.append(Paragraph('Executive Summary', self.styles['SectionHeader']))
        
        summary_data = [
            ['Metric', 'Value'],
            ['Quality Score', f"{results.get('score', 100)}%"],
            ['Grade', results.get('grade', 'A')],
            ['Total Issues', str(results.get('issue_count', 0))],
        ]
        
        doc_info = results.get('document_info', {})
        if doc_info:
            summary_data.extend([
                ['Word Count', str(doc_info.get('word_count', 'N/A'))],
                ['Paragraphs', str(doc_info.get('paragraph_count', 'N/A'))],
            ])
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DEE2E6')),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Severity breakdown
        story.append(Paragraph('Issues by Severity', self.styles['SectionHeader']))
        
        by_severity = results.get('by_severity', {})
        severity_data = [['Severity', 'Count']]
        severity_colors = {
            'Critical': '#DC3545',
            'High': '#FD7E14',
            'Medium': '#FFC107',
            'Low': '#28A745',
            'Info': '#17A2B8'
        }
        
        for sev in ['Critical', 'High', 'Medium', 'Low', 'Info']:
            severity_data.append([sev, str(by_severity.get(sev, 0))])
        
        sev_table = Table(severity_data, colWidths=[2*inch, 1.5*inch])
        sev_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#DEE2E6')),
        ]))
        story.append(sev_table)
        story.append(Spacer(1, 20))
        
        # Top issues
        story.append(Paragraph('Top Issues', self.styles['SectionHeader']))
        
        issues = results.get('issues', [])[:20]  # Top 20
        if issues:
            issue_data = [['#', 'Severity', 'Category', 'Message']]
            for i, issue in enumerate(issues, 1):
                msg = issue.get('message', '')[:60] + ('...' if len(issue.get('message', '')) > 60 else '')
                issue_data.append([
                    str(i),
                    issue.get('severity', ''),
                    issue.get('category', ''),
                    msg
                ])
            
            issue_table = Table(issue_data, colWidths=[0.4*inch, 0.8*inch, 1.5*inch, 4*inch])
            issue_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A5F')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#DEE2E6')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ]))
            story.append(issue_table)
        
        # Build PDF
        doc.build(story)
        
        pdf_content = output.getvalue()
        
        if filename:
            with open(filename, 'wb') as f:
                f.write(pdf_content)
        
        return pdf_content


class JSONExporter:
    """Export analysis results to JSON format.
    
    v2.2.0: Added provenance summary statistics
    """
    
    @staticmethod
    def export(results: Dict, filename: str = None, pretty: bool = True, 
               include_provenance_summary: bool = True) -> str:
        """Export results to JSON.
        
        Args:
            results: Analysis results dictionary
            filename: Optional output file path
            pretty: Whether to format with indentation
            include_provenance_summary: Whether to add provenance statistics
        """
        # Clean up results for JSON serialization
        clean_results = JSONExporter._clean_for_json(results)
        
        # Add provenance summary if requested and issues exist
        if include_provenance_summary and 'issues' in clean_results:
            clean_results['provenance_summary'] = JSONExporter._compute_provenance_summary(
                clean_results.get('issues', [])
            )
        
        json_content = json.dumps(clean_results, indent=2 if pretty else None, default=str)
        
        if filename:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(json_content)
        
        return json_content
    
    @staticmethod
    def _compute_provenance_summary(issues: List[Dict]) -> Dict:
        """Compute provenance tracking statistics from issues.
        
        Returns summary with:
        - total_issues: Total number of issues
        - validated_count: Issues with is_validated=True
        - unvalidated_count: Issues without provenance or is_validated=False
        - validation_rate: Percentage of issues validated
        - by_category: Breakdown by issue category
        """
        total = len(issues)
        validated = 0
        unvalidated = 0
        by_category = defaultdict(lambda: {'validated': 0, 'unvalidated': 0, 'total': 0})
        
        for issue in issues:
            category = issue.get('category', 'Unknown')
            source = issue.get('source', {})
            is_validated = source.get('is_validated', False) if isinstance(source, dict) else False
            
            by_category[category]['total'] += 1
            
            if is_validated:
                validated += 1
                by_category[category]['validated'] += 1
            else:
                unvalidated += 1
                by_category[category]['unvalidated'] += 1
        
        # Compute rates for each category
        category_summary = {}
        for cat, counts in by_category.items():
            rate = (counts['validated'] / counts['total'] * 100) if counts['total'] > 0 else 0
            category_summary[cat] = {
                'validated': counts['validated'],
                'unvalidated': counts['unvalidated'],
                'total': counts['total'],
                'validation_rate': round(rate, 1)
            }
        
        return {
            'total_issues': total,
            'validated_count': validated,
            'unvalidated_count': unvalidated,
            'validation_rate': round((validated / total * 100) if total > 0 else 0, 1),
            'by_category': category_summary,
            'note': 'Validated issues have been confirmed to exist in the original document text.'
        }
    
    @staticmethod
    def _clean_for_json(obj: Any) -> Any:
        """Clean object for JSON serialization."""
        if isinstance(obj, dict):
            return {k: JSONExporter._clean_for_json(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [JSONExporter._clean_for_json(v) for v in obj]
        elif isinstance(obj, set):
            return list(obj)
        elif hasattr(obj, '__dict__'):
            return JSONExporter._clean_for_json(obj.__dict__)
        else:
            return obj


class ComplianceMatrixExporter:
    """Export compliance matrix mapping issues to standards."""
    
    STANDARDS = {
        'MIL-STD-961': {
            'passive_voice': '5.1.2 - Clear Requirements',
            'weak_language': '5.1.3 - Definitive Language',
            'ambiguous_pronouns': '5.1.4 - Unambiguous',
            'testability': '5.1.5 - Verifiable',
            'atomicity': '5.1.6 - Single Requirement',
        },
        'AS6500': {
            'requirements_language': '4.2 - Design Analysis',
            'roles': '4.5.8 - Supplier Management',
            'acronyms': '2.2 - Definitions',
        },
        'DO-178C': {
            'testability': '6.3 - Verification',
            'ambiguous_pronouns': '6.2 - Clarity',
            'tbd': '6.1 - Completeness',
        }
    }
    
    @staticmethod
    def generate_matrix(issues: List[Dict], standard: str = None) -> Dict:
        """Generate compliance matrix."""
        matrix = {}
        
        standards_to_check = [standard] if standard else ComplianceMatrixExporter.STANDARDS.keys()
        
        for std in standards_to_check:
            if std not in ComplianceMatrixExporter.STANDARDS:
                continue
            
            std_mapping = ComplianceMatrixExporter.STANDARDS[std]
            matrix[std] = {
                'clauses': [],
                'total_issues': 0
            }
            
            for issue in issues:
                category = issue.get('category', '').lower().replace(' ', '_')
                if category in std_mapping:
                    clause = std_mapping[category]
                    existing = next((c for c in matrix[std]['clauses'] if c['clause'] == clause), None)
                    if existing:
                        existing['count'] += 1
                        existing['issues'].append(issue)
                    else:
                        matrix[std]['clauses'].append({
                            'clause': clause,
                            'category': issue.get('category'),
                            'count': 1,
                            'issues': [issue]
                        })
                    matrix[std]['total_issues'] += 1
        
        return matrix


# Factory function
def get_exporter(format_type: str):
    """Get appropriate exporter for format type."""
    exporters = {
        'excel': ExcelExporter,
        'xlsx': ExcelExporter,
        'csv': CSVExporter,
        'pdf': PDFExporter,
        'json': JSONExporter,
    }
    
    exporter_class = exporters.get(format_type.lower())
    if not exporter_class:
        raise ValueError(f"Unsupported export format: {format_type}")
    
    return exporter_class()


# v3.0.33 Chunk D: Helper functions for enhanced export
def generate_timestamped_filename(base_name: str, extension: str = 'xlsx') -> str:
    """Generate a filename with timestamp.
    
    Args:
        base_name: Base filename without extension (e.g., 'review_export')
        extension: File extension without dot (e.g., 'xlsx')
    
    Returns:
        Filename with timestamp, e.g., 'review_export_20260120_153045.xlsx'
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f"{base_name}_{timestamp}.{extension}"


def export_xlsx_enhanced(results: Dict, 
                         base_filename: str = 'review_export',
                         severities: List[str] = None,
                         document_metadata: Dict = None) -> tuple:
    """Enhanced XLSX export with all Chunk D features.
    
    v3.0.33 Chunk D: Convenience function that combines:
    - Timestamped filename generation
    - Severity filtering
    - Document metadata inclusion
    - Action Item column
    
    Args:
        results: Analysis results dictionary
        base_filename: Base name for the file
        severities: Optional list of severities to filter
        document_metadata: Optional dict with filename, scan_date, score
    
    Returns:
        tuple: (filename with timestamp, bytes content)
    """
    exporter = ExcelExporter()
    filename = generate_timestamped_filename(base_filename, 'xlsx')
    content = exporter.export(
        results,
        severities=severities,
        document_metadata=document_metadata
    )
    return filename, content
