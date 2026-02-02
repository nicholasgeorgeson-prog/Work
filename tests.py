#!/usr/bin/env python3
"""
TechWriterReview Test Suite v3.0.103
====================================
Validates security controls, API endpoints, and core functionality.

Run with: python -m pytest tests.py -v
Or standalone: python tests.py

Changelog:
- v3.0.103: Added Fix Assistant v2 API tests, batch limit tests, docstrings
- v2.8.0: Initial comprehensive test suite
"""

import os
import sys
import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch, MagicMock
from io import BytesIO

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent))

# Import application
from app import app, config, SessionManager
from config_logging import (
    get_config, VERSION, RateLimiter,
    ValidationError, TechWriterError,
    sanitize_filename, validate_file_extension,
    reset_config, reset_rate_limiter
)


class TestSecurityControls(unittest.TestCase):
    """Test security-related functionality."""
    
    def setUp(self):
        """Set up test client."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
    
    def test_csrf_token_endpoint(self):
        """
        Test CSRF token endpoint returns valid token.
        
        Security: CSRF protection prevents cross-site request forgery.
        Expects: 200 response with csrf_token field (min 32 chars).
        """
        response = self.client.get('/api/csrf-token')
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertIn('csrf_token', data)
        self.assertTrue(len(data['csrf_token']) >= 32)
    
    def test_file_size_limit_configured(self):
        """
        Test that file size limit is configured.
        
        Security: File size limits prevent denial of service attacks.
        Expects: MAX_CONTENT_LENGTH is set between 0 and 100MB.
        """
        self.assertIsNotNone(app.config.get('MAX_CONTENT_LENGTH'))
        self.assertGreater(app.config['MAX_CONTENT_LENGTH'], 0)
        self.assertLessEqual(app.config['MAX_CONTENT_LENGTH'], 100 * 1024 * 1024)
    
    def test_security_headers_present(self):
        """
        Test that security headers are added to responses.
        
        Security: Headers prevent XSS, clickjacking, and MIME sniffing attacks.
        Expects: X-Content-Type-Options, X-Frame-Options, X-XSS-Protection present.
        """
        response = self.client.get('/api/health')
        
        self.assertIn('X-Content-Type-Options', response.headers)
        self.assertEqual(response.headers['X-Content-Type-Options'], 'nosniff')
        
        self.assertIn('X-Frame-Options', response.headers)
        self.assertEqual(response.headers['X-Frame-Options'], 'DENY')
        
        self.assertIn('X-XSS-Protection', response.headers)
        self.assertIn('Referrer-Policy', response.headers)
    
    def test_rate_limiter_blocks_excess_requests(self):
        """
        Test rate limiter blocks requests over limit.
        
        Security: Rate limiting prevents brute force and DoS attacks.
        Expects: Requests beyond limit are blocked (returns False).
        """
        limiter = RateLimiter(max_requests=3, window_seconds=60)
        
        for i in range(3):
            self.assertTrue(limiter.is_allowed('test_ip'))
        
        self.assertFalse(limiter.is_allowed('test_ip'))
    
    def test_rate_limiter_retry_after(self):
        """
        Test rate limiter returns retry_after value.
        
        Security: Informs clients when they can retry after being rate-limited.
        Expects: retry_after > 0 and <= window_seconds when blocked.
        """
        limiter = RateLimiter(max_requests=1, window_seconds=60)
        limiter.is_allowed('test_ip')
        limiter.is_allowed('test_ip')  # Blocked
        
        retry_after = limiter.get_retry_after('test_ip')
        self.assertGreater(retry_after, 0)
        self.assertLessEqual(retry_after, 60)


class TestFileValidation(unittest.TestCase):
    """Test file upload validation."""
    
    def test_sanitize_filename_removes_path_separators(self):
        """
        Test filename sanitization removes dangerous characters.
        
        Security: Prevents path traversal attacks via malicious filenames.
        Expects: No /, \\, null bytes, or leading dots in sanitized output.
        """
        dangerous_names = [
            '../../../etc/passwd',
            '..\\..\\windows\\system32',
            'file\x00.docx',
            '.hidden',
        ]
        
        for name in dangerous_names:
            sanitized = sanitize_filename(name)
            self.assertNotIn('/', sanitized)
            self.assertNotIn('\\', sanitized)
            self.assertNotIn('\x00', sanitized)
            self.assertFalse(sanitized.startswith('.'))
    
    def test_sanitize_filename_preserves_valid_names(self):
        """
        Test valid filenames are preserved.
        
        Functionality: Normal filenames should pass through unchanged.
        Expects: Output has length > 0 and retains original extension.
        """
        valid_names = [
            'document.docx',
            'my-file_v2.pdf',
            'Report 2024.docx',
        ]
        
        for name in valid_names:
            sanitized = sanitize_filename(name)
            self.assertTrue(len(sanitized) > 0)
            self.assertTrue(sanitized.endswith('.docx') or sanitized.endswith('.pdf'))
    
    def test_validate_file_extension_allows_valid(self):
        """
        Test valid file extensions are accepted.
        
        Functionality: .docx and .pdf should be allowed (case-insensitive).
        Expects: validate_file_extension returns True for valid types.
        """
        self.assertTrue(validate_file_extension('document.docx'))
        self.assertTrue(validate_file_extension('document.DOCX'))
        self.assertTrue(validate_file_extension('document.pdf'))
        self.assertTrue(validate_file_extension('document.PDF'))
    
    def test_validate_file_extension_rejects_invalid(self):
        """
        Test invalid file extensions are rejected.
        
        Security: Only allow known safe document types.
        Expects: validate_file_extension returns False for dangerous types.
        """
        self.assertFalse(validate_file_extension('document.exe'))
        self.assertFalse(validate_file_extension('document.php'))
        self.assertFalse(validate_file_extension('document.js'))
        self.assertFalse(validate_file_extension('document.html'))


class TestAPIEndpoints(unittest.TestCase):
    """Test API endpoint functionality."""
    
    def setUp(self):
        """Set up test client."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
    
    def test_health_endpoint(self):
        """
        Test health endpoint returns healthy status.
        
        Monitoring: Health endpoint is used by load balancers/orchestrators.
        Expects: 200 response with status='healthy', version, and timestamp.
        """
        response = self.client.get('/api/health')
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertEqual(data['status'], 'healthy')
        self.assertIn('version', data)
        self.assertIn('timestamp', data)
    
    def test_ready_endpoint(self):
        """
        Test readiness endpoint.
        
        Monitoring: Readiness checks verify all dependencies are available.
        Expects: 200 (ready) or 503 (not ready) with ready flag and checks.
        """
        response = self.client.get('/api/ready')
        self.assertIn(response.status_code, [200, 503])
        data = json.loads(response.data)
        self.assertIn('ready', data)
        self.assertIn('checks', data)
    
    def test_version_endpoint(self):
        """
        Test version endpoint returns correct info.
        
        Functionality: UI version label depends on this endpoint.
        Expects: 200 response with app_version matching VERSION constant.
        """
        response = self.client.get('/api/version')
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertEqual(data['app_version'], VERSION)
        self.assertIn('core_version', data)
        self.assertIn('api_version', data)
    
    def test_config_get_endpoint(self):
        """
        Test config GET endpoint (user configuration).
        
        Functionality: Returns user preferences for UI configuration.
        Expects: 200 response with success=True and data field.
        
        Note: /api/userconfig is handled by api_extensions for user preferences.
        /api/config/sharing is handled by app.py for sharing settings.
        """
        response = self.client.get('/api/userconfig')
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertTrue(data['success'])
        self.assertIn('data', data)
    
    def test_main_config_endpoint(self):
        """
        Test main /api/config endpoint uses standardized 'data' envelope.
        
        API Design: Consistent response format across all endpoints.
        Expects: success=True, uses 'data' key (not legacy 'config' key).
        """
        response = self.client.get('/api/config')
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertTrue(data['success'])
        # Verify standardized envelope (data, not config)
        self.assertIn('data', data)
        self.assertNotIn('config', data)  # Old envelope key should not exist


class TestErrorHandling(unittest.TestCase):
    """Test error handling and responses."""
    
    def setUp(self):
        """Set up test client."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
    
    def test_404_returns_json(self):
        """
        Test 404 errors return JSON response.
        
        API Design: All errors should be machine-readable JSON.
        Expects: 404 status with success=False and error message.
        """
        response = self.client.get('/api/nonexistent')
        self.assertEqual(response.status_code, 404)
        data = json.loads(response.data)
        self.assertFalse(data['success'])
        self.assertIn('error', data)
    
    def test_validation_error_structure(self):
        """
        Test ValidationError has correct structure.
        
        API Design: Custom error classes provide consistent error format.
        Expects: status_code=400, code=VALIDATION_ERROR, proper to_dict().
        """
        err = ValidationError("Test error", field="test_field")
        self.assertEqual(err.status_code, 400)
        self.assertEqual(err.code, "VALIDATION_ERROR")
        
        error_dict = err.to_dict()
        self.assertFalse(error_dict['success'])
        self.assertIn('error', error_dict)
        self.assertEqual(error_dict['error']['code'], 'VALIDATION_ERROR')


class TestSessionManagement(unittest.TestCase):
    """Test session management functionality."""
    
    def test_session_create(self):
        """
        Test session creation.
        
        Functionality: Sessions store user state across requests.
        Expects: create() returns session_id, get() returns data with 'created'.
        """
        session_id = SessionManager.create()
        self.assertIsNotNone(session_id)
        
        session_data = SessionManager.get(session_id)
        self.assertIsNotNone(session_data)
        self.assertIn('created', session_data)
    
    def test_session_update(self):
        """
        Test session update.
        
        Functionality: Sessions can be updated with new data.
        Expects: update() stores data retrievable by get().
        """
        session_id = SessionManager.create()
        SessionManager.update(session_id, current_file='test.docx')
        
        session_data = SessionManager.get(session_id)
        self.assertEqual(session_data['current_file'], 'test.docx')
    
    def test_session_delete(self):
        """
        Test session deletion.
        
        Functionality: Sessions can be explicitly deleted.
        Expects: delete() removes session, get() returns None afterward.
        """
        session_id = SessionManager.create()
        self.assertIsNotNone(SessionManager.get(session_id))
        
        SessionManager.delete(session_id)
        self.assertIsNone(SessionManager.get(session_id))
    
    def test_session_get_nonexistent(self):
        """
        Test getting nonexistent session returns None.
        
        Error Handling: Querying invalid session should not raise exception.
        Expects: get() returns None for nonexistent session ID.
        """
        self.assertIsNone(SessionManager.get('nonexistent-id'))


class TestVersionConsistency(unittest.TestCase):
    """Test version consistency across modules."""
    
    def test_version_string_format(self):
        """
        Test version string is properly formatted.
        
        Consistency: Version must follow semantic versioning (X.Y.Z).
        Expects: Three numeric parts separated by dots.
        """
        from config_logging import VERSION
        parts = VERSION.split('.')
        self.assertEqual(len(parts), 3)
        for part in parts:
            self.assertTrue(part.isdigit())
    
    def test_version_matches_json(self):
        """
        Test version matches version.json.
        
        Consistency: config_logging.VERSION must match version.json.
        Expects: VERSION equals version field in version.json file.
        """
        from config_logging import VERSION
        import json
        from pathlib import Path
        version_file = Path(__file__).parent / 'version.json'
        if version_file.exists():
            with open(version_file, 'r') as f:
                version_data = json.load(f)
            self.assertEqual(VERSION, version_data.get('version', VERSION))
    
    def test_core_version_matches(self):
        """
        Test core module version matches config.
        
        Consistency: core.py MODULE_VERSION must match config_logging VERSION.
        Expects: Both versions are identical strings.
        """
        from config_logging import VERSION
        from core import MODULE_VERSION
        self.assertEqual(VERSION, MODULE_VERSION)


class TestCodeQuality(unittest.TestCase):
    """Static code quality checks."""
    
    def test_no_bare_except_in_app(self):
        """
        Test app.py has no bare except clauses.
        
        Code Quality: Bare excepts catch too much, hiding bugs.
        Expects: Zero matches for '^\\s*except:\\s*$' in app.py.
        """
        app_path = Path(__file__).parent / 'app.py'
        if app_path.exists():
            with open(app_path, 'r') as f:
                content = f.read()
            
            import re
            bare_excepts = re.findall(r'^\s*except:\s*$', content, re.MULTILINE)
            self.assertEqual(len(bare_excepts), 0, 
                f"Found {len(bare_excepts)} bare 'except:' clauses")
    
    def test_no_bare_except_in_config(self):
        """
        Test config_logging.py has no bare except clauses.
        
        Code Quality: Bare excepts should be replaced with specific exceptions.
        Expects: Zero bare except clauses in config_logging.py.
        """
        config_path = Path(__file__).parent / 'config_logging.py'
        if config_path.exists():
            with open(config_path, 'r') as f:
                content = f.read()
            
            import re
            bare_excepts = re.findall(r'^\s*except:\s*$', content, re.MULTILINE)
            self.assertEqual(len(bare_excepts), 0)


class TestConfigDefaults(unittest.TestCase):
    """Test configuration defaults."""
    
    def test_debug_default_false(self):
        """
        Test debug defaults to False.
        
        Security: Debug mode exposes sensitive information in production.
        Expects: get_config().debug is False by default.
        """
        cfg = get_config()
        self.assertFalse(cfg.debug)
    
    def test_csrf_default_enabled(self):
        """
        Test CSRF defaults to enabled.
        
        Security: CSRF protection must be on by default.
        Expects: get_config().csrf_enabled is True.
        """
        cfg = get_config()
        self.assertTrue(cfg.csrf_enabled)
    
    def test_rate_limit_default_enabled(self):
        """
        Test rate limiting defaults to enabled.
        
        Security: Rate limiting must be on by default.
        Expects: get_config().rate_limit_enabled is True.
        """
        cfg = get_config()
        self.assertTrue(cfg.rate_limit_enabled)
    
    def test_allowed_extensions_includes_pdf(self):
        """
        Test PDF is in allowed extensions.
        
        Functionality: PDF and DOCX are core supported formats.
        Expects: Both .pdf and .docx in allowed_extensions.
        """
        cfg = get_config()
        self.assertIn('.pdf', cfg.allowed_extensions)
        self.assertIn('.docx', cfg.allowed_extensions)


class TestRateLimitIntegration(unittest.TestCase):
    """Integration tests for rate limiting returning 429."""
    
    def setUp(self):
        """Set up test client with rate limiting."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
        # Reset rate limiter for clean test
        reset_rate_limiter()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
        reset_rate_limiter()
    
    def test_rate_limit_returns_429_on_excess(self):
        """
        Test that rate limiting returns 429 when limit exceeded on non-exempt endpoints.
        
        Security: Rate-limited requests must receive 429 Too Many Requests.
        Expects: Third request to /api/config returns 429 with RATE_LIMIT code.
        """
        # Create a limiter with very low limit for testing
        from config_logging import _rate_limiter, RateLimiter
        import config_logging
        
        # Save original and set test limiter
        original_limiter = config_logging._rate_limiter
        test_limiter = RateLimiter(max_requests=2, window_seconds=60)
        config_logging._rate_limiter = test_limiter
        
        try:
            # Use /api/config which is NOT exempt from rate limiting
            # First 2 requests should succeed
            response1 = self.client.get('/api/config')
            response2 = self.client.get('/api/config')
            
            # Third request should be rate limited (429)
            response3 = self.client.get('/api/config')
            
            # Check that we got a 429
            self.assertEqual(response3.status_code, 429)
            data = json.loads(response3.data)
            self.assertFalse(data['success'])
            self.assertEqual(data['error']['code'], 'RATE_LIMIT')
            self.assertIn('retry_after', data['error'])
        finally:
            # Restore original limiter
            config_logging._rate_limiter = original_limiter
    
    def test_rate_limit_exempt_endpoints(self):
        """
        Test that health/status endpoints are exempt from rate limiting.
        
        Monitoring: Health checks must always succeed for orchestrators.
        Expects: Multiple requests to /api/health all return 200.
        """
        from config_logging import RateLimiter
        import config_logging
        
        # Save original and set very restrictive test limiter
        original_limiter = config_logging._rate_limiter
        test_limiter = RateLimiter(max_requests=1, window_seconds=60)
        config_logging._rate_limiter = test_limiter
        
        try:
            # Exempt endpoints should all succeed even with max_requests=1
            # These are exempt: /api/health, /api/ready, /api/version, /api/csrf-token
            
            response1 = self.client.get('/api/health')
            self.assertEqual(response1.status_code, 200)
            
            response2 = self.client.get('/api/health')
            self.assertEqual(response2.status_code, 200)
            
            response3 = self.client.get('/api/health')
            self.assertEqual(response3.status_code, 200)
            
            # All should be 200, not 429
            self.assertNotEqual(response1.status_code, 429)
            self.assertNotEqual(response2.status_code, 429)
            self.assertNotEqual(response3.status_code, 429)
        finally:
            # Restore original limiter
            config_logging._rate_limiter = original_limiter


class TestAuthenticationIntegration(unittest.TestCase):
    """Integration tests for authentication returning 401."""
    
    def setUp(self):
        """Set up test client."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
        reset_config()
    
    def test_auth_returns_401_when_enabled_without_header(self):
        """
        Test that auth returns 401 when enabled but header missing.
        
        Security: Protected endpoints require authentication.
        Expects: 401 response with AUTH_REQUIRED code when header missing.
        """
        import config_logging
        import app as app_module
        
        # Save original config
        original_config = config_logging._config
        original_app_config = app_module.config
        
        try:
            # Create config with auth enabled (trusted_header mode)
            os.environ['TWR_AUTH'] = 'true'
            os.environ['TWR_AUTH_PROVIDER'] = 'trusted_header'
            
            # Force reload config in both places
            config_logging._config = None
            new_config = config_logging.get_config()
            app_module.config = new_config
            
            # Verify auth is enabled
            self.assertTrue(new_config.auth_enabled)
            self.assertEqual(new_config.auth_provider, 'trusted_header')
            
            # Request to protected endpoint without auth header should return 401
            # Note: /api/version is exempt from auth, so we use /api/config
            response = self.client.get('/api/config')
            self.assertEqual(response.status_code, 401)
            data = json.loads(response.data)
            self.assertFalse(data['success'])
            self.assertEqual(data['error']['code'], 'AUTH_REQUIRED')
        finally:
            # Restore original config
            os.environ.pop('TWR_AUTH', None)
            os.environ.pop('TWR_AUTH_PROVIDER', None)
            config_logging._config = original_config
            app_module.config = original_app_config
    
    def test_auth_allows_request_with_valid_header(self):
        """
        Test that auth allows request when valid header provided.
        
        Functionality: Properly authenticated requests should succeed.
        Expects: 200 response when X-Authenticated-User header present.
        """
        import config_logging
        import app as app_module
        
        # Save original config
        original_config = config_logging._config
        original_app_config = app_module.config
        
        try:
            # Create config with auth enabled (trusted_header mode)
            os.environ['TWR_AUTH'] = 'true'
            os.environ['TWR_AUTH_PROVIDER'] = 'trusted_header'
            
            # Force reload config
            config_logging._config = None
            new_config = config_logging.get_config()
            app_module.config = new_config
            
            # Request with auth header should succeed
            response = self.client.get('/api/version', 
                headers={'X-Authenticated-User': 'testuser'})
            self.assertEqual(response.status_code, 200)
        finally:
            # Restore original config
            os.environ.pop('TWR_AUTH', None)
            os.environ.pop('TWR_AUTH_PROVIDER', None)
            config_logging._config = original_config
            app_module.config = original_app_config
    
    def test_health_endpoint_exempt_from_auth(self):
        """
        Test that /api/health is exempt from authentication.
        
        Monitoring: Health checks must work regardless of auth state.
        Expects: 200 response even when auth enabled without header.
        """
        import config_logging
        import app as app_module
        
        # Save original config
        original_config = config_logging._config
        original_app_config = app_module.config
        
        try:
            # Create config with auth enabled
            os.environ['TWR_AUTH'] = 'true'
            os.environ['TWR_AUTH_PROVIDER'] = 'trusted_header'
            
            # Force reload config
            config_logging._config = None
            new_config = config_logging.get_config()
            app_module.config = new_config
            
            # Health endpoint should work without auth
            response = self.client.get('/api/health')
            self.assertEqual(response.status_code, 200)
        finally:
            # Restore original config
            os.environ.pop('TWR_AUTH', None)
            os.environ.pop('TWR_AUTH_PROVIDER', None)
            config_logging._config = original_config
            app_module.config = original_app_config
    
    def test_version_endpoint_exempt_from_auth(self):
        """
        Test that /api/version is exempt from authentication.
        
        Functionality: UI version label must work without authentication.
        Expects: 200 response with version data even when auth enabled.
        """
        import config_logging
        import app as app_module
        
        # Save original config
        original_config = config_logging._config
        original_app_config = app_module.config
        
        try:
            # Create config with auth enabled
            os.environ['TWR_AUTH'] = 'true'
            os.environ['TWR_AUTH_PROVIDER'] = 'trusted_header'
            
            # Force reload config
            config_logging._config = None
            new_config = config_logging.get_config()
            app_module.config = new_config
            
            # Version endpoint should work without auth (for UI version label)
            response = self.client.get('/api/version')
            self.assertEqual(response.status_code, 200)
            
            # Verify it returns version data
            data = json.loads(response.data)
            self.assertIn('app_version', data)
        finally:
            # Restore original config
            os.environ.pop('TWR_AUTH', None)
            os.environ.pop('TWR_AUTH_PROVIDER', None)
            config_logging._config = original_config
            app_module.config = original_app_config


class TestAPIKeyAuth(unittest.TestCase):
    """Test API key authentication mode."""
    
    def setUp(self):
        """Set up test client."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
        reset_config()
    
    def test_api_key_auth_rejects_invalid_key(self):
        """
        Test that API key auth rejects invalid keys.
        
        Security: Invalid API keys must be rejected.
        Expects: 401 response when X-API-Key header has wrong value.
        """
        import config_logging
        import app as app_module
        
        original_config = config_logging._config
        original_app_config = app_module.config
        
        try:
            os.environ['TWR_AUTH'] = 'true'
            os.environ['TWR_AUTH_PROVIDER'] = 'api_key'
            os.environ['TWR_API_KEY'] = 'secret-test-key-12345'
            
            config_logging._config = None
            new_config = config_logging.get_config()
            app_module.config = new_config
            
            # Request with wrong key to non-exempt endpoint
            # Note: /api/version is exempt from auth, so we use /api/config
            response = self.client.get('/api/config',
                headers={'X-API-Key': 'wrong-key'})
            self.assertEqual(response.status_code, 401)
        finally:
            os.environ.pop('TWR_AUTH', None)
            os.environ.pop('TWR_AUTH_PROVIDER', None)
            os.environ.pop('TWR_API_KEY', None)
            config_logging._config = original_config
            app_module.config = original_app_config
    
    def test_api_key_auth_accepts_valid_key(self):
        """
        Test that API key auth accepts valid keys.
        
        Functionality: Valid API key should grant access.
        Expects: 200 response when X-API-Key header matches configured key.
        """
        import config_logging
        import app as app_module
        
        original_config = config_logging._config
        original_app_config = app_module.config
        
        try:
            os.environ['TWR_AUTH'] = 'true'
            os.environ['TWR_AUTH_PROVIDER'] = 'api_key'
            os.environ['TWR_API_KEY'] = 'secret-test-key-12345'
            
            config_logging._config = None
            new_config = config_logging.get_config()
            app_module.config = new_config
            
            # Request with correct key to non-exempt endpoint
            # Note: /api/version is exempt from auth, so we use /api/config
            response = self.client.get('/api/config',
                headers={'X-API-Key': 'secret-test-key-12345'})
            self.assertEqual(response.status_code, 200)
        finally:
            os.environ.pop('TWR_AUTH', None)
            os.environ.pop('TWR_AUTH_PROVIDER', None)
            os.environ.pop('TWR_API_KEY', None)
            config_logging._config = original_config
            app_module.config = original_app_config


class TestRolesGraphAPI(unittest.TestCase):
    """Test the /api/roles/graph endpoint."""
    
    def setUp(self):
        """Set up test client."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
    
    def test_roles_graph_endpoint_exists(self):
        """
        Test that /api/roles/graph endpoint exists and returns JSON.
        
        Functionality: Roles graph visualizes role relationships.
        Expects: 200 or 500 (if no scan history), response contains 'success'.
        """
        response = self.client.get('/api/roles/graph')
        self.assertIn(response.status_code, [200, 500])  # 500 if scan history unavailable
        data = json.loads(response.data)
        self.assertIn('success', data)
    
    def test_roles_graph_response_schema(self):
        """
        Test that graph response has correct schema.
        
        API Design: Graph data must have nodes, links, and count mappings.
        Expects: data contains nodes, links, role_counts, doc_counts arrays/dicts.
        """
        response = self.client.get('/api/roles/graph')
        data = json.loads(response.data)
        
        if data.get('success'):
            self.assertIn('data', data)
            graph_data = data['data']
            
            # Check required keys
            self.assertIn('nodes', graph_data)
            self.assertIn('links', graph_data)
            self.assertIn('role_counts', graph_data)
            self.assertIn('doc_counts', graph_data)
            
            # Nodes should be a list
            self.assertIsInstance(graph_data['nodes'], list)
            self.assertIsInstance(graph_data['links'], list)
            
            # If there are nodes, check their structure
            if len(graph_data['nodes']) > 0:
                node = graph_data['nodes'][0]
                self.assertIn('id', node)
                self.assertIn('label', node)
                self.assertIn('type', node)
                
                # IDs should be stable format
                self.assertTrue(
                    node['id'].startswith('role_') or node['id'].startswith('doc_'),
                    f"Node ID should have stable prefix: {node['id']}"
                )
    
    def test_roles_graph_query_params(self):
        """
        Test that query parameters are respected.
        
        Functionality: Graph supports filtering via max_nodes and min_weight.
        Expects: Response metadata reflects requested parameter values.
        """
        # Test max_nodes
        response = self.client.get('/api/roles/graph?max_nodes=10')
        data = json.loads(response.data)
        if data.get('success') and data.get('data', {}).get('meta'):
            self.assertEqual(data['data']['meta']['max_nodes'], 10)
        
        # Test min_weight
        response = self.client.get('/api/roles/graph?min_weight=5')
        data = json.loads(response.data)
        if data.get('success') and data.get('data', {}).get('meta'):
            self.assertEqual(data['data']['meta']['min_weight'], 5)
    
    def test_roles_graph_max_nodes_limit(self):
        """
        Test that max_nodes is capped at 500.
        
        Performance: Large graphs are limited to prevent browser crashes.
        Expects: max_nodes in response is <= 500 even if 1000 requested.
        """
        response = self.client.get('/api/roles/graph?max_nodes=1000')
        data = json.loads(response.data)
        if data.get('success') and data.get('data', {}).get('meta'):
            # Should be capped at 500
            self.assertLessEqual(data['data']['meta']['max_nodes'], 500)


class TestCapabilitiesAPI(unittest.TestCase):
    """Test the /api/capabilities endpoint."""
    
    def setUp(self):
        """Set up test client."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
    
    def test_capabilities_endpoint_exists(self):
        """
        Test that /api/capabilities endpoint exists.
        
        Functionality: UI uses capabilities to enable/disable features.
        Expects: 200 response with success=True.
        """
        response = self.client.get('/api/capabilities')
        self.assertEqual(response.status_code, 200)
        data = json.loads(response.data)
        self.assertTrue(data.get('success'))
    
    def test_capabilities_response_schema(self):
        """
        Test capabilities response has correct schema.
        
        API Design: Capabilities endpoint structure for feature detection.
        Expects: data contains version and capabilities dict with feature flags.
        
        The /api/capabilities endpoint is now handled by api_extensions.py
        and returns: { success, data: { version, capabilities: {...} } }
        """
        response = self.client.get('/api/capabilities')
        data = json.loads(response.data)
        
        self.assertIn('data', data)
        self.assertIn('version', data['data'])
        self.assertIn('capabilities', data['data'])
        
        caps = data['data']['capabilities']
        self.assertIsInstance(caps, dict)
        
        # Check for core feature capabilities
        self.assertIn('roles_graph', caps)
        self.assertIsInstance(caps['roles_graph'], bool)
        
        # Check for export capabilities (used by frontend UI)
        self.assertIn('excel_export', caps)
        self.assertIn('pdf_export', caps)
        self.assertIsInstance(caps['excel_export'], bool)
        self.assertIsInstance(caps['pdf_export'], bool)


class TestStaticFileSecurity(unittest.TestCase):
    """
    Test security of static file serving routes.
    
    Ensures that static routes cannot be used for file disclosure attacks.
    """
    
    def setUp(self):
        """Set up test client."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
    
    def test_images_rejects_non_image_extensions(self):
        """
        Test that /static/images/* rejects non-image file requests.
        
        Security: Prevent file disclosure via static image routes.
        Expects: 404 for non-image extensions like .json, .py, .db.
        """
        # These should all return 404, not serve files
        non_image_files = [
            'config.json',
            'app.py',
            'database.db',
            'secrets.txt',
            '../config.json',
            '..%2Fconfig.json',
        ]
        
        for filename in non_image_files:
            response = self.client.get(f'/static/images/{filename}')
            self.assertEqual(response.status_code, 404,
                f"Should reject non-image file: {filename}")
    
    def test_images_allows_valid_image_extensions(self):
        """
        Test that /static/images/* allows valid image extensions.
        
        Functionality: Valid image files should be servable.
        Expects: 200 or 404 (file not found), not 500 or blocked status.
        
        Note: Returns 404 if file doesn't exist, but should not reject based on extension.
        """
        valid_extensions = ['test.png', 'test.jpg', 'test.jpeg', 'test.gif', 
                           'test.svg', 'test.ico', 'test.webp']
        
        for filename in valid_extensions:
            response = self.client.get(f'/static/images/{filename}')
            # Should be 404 (file not found) not blocked by extension check
            # The important thing is it's not a 500 or serving wrong content
            self.assertIn(response.status_code, [200, 404],
                f"Valid image extension should not error: {filename}")
    
    def test_css_rejects_non_css_extensions(self):
        """
        Test that /static/css/* only serves .css files.
        
        Security: CSS route should not serve non-CSS files.
        Expects: 400 or 404 for .json and .py files in CSS directory.
        """
        response = self.client.get('/static/css/config.json')
        self.assertIn(response.status_code, [400, 404])
        
        response = self.client.get('/static/css/app.py')
        self.assertIn(response.status_code, [400, 404])
    
    def test_js_rejects_non_js_extensions(self):
        """
        Test that /static/js/* only serves .js files.
        
        Security: JS route should not serve non-JS files.
        Expects: 400 or 404 for .json and .py files in JS directory.
        """
        response = self.client.get('/static/js/config.json')
        self.assertIn(response.status_code, [400, 404])
        
        response = self.client.get('/static/js/app.py')
        self.assertIn(response.status_code, [400, 404])
    
    def test_vendor_rejects_non_js_extensions(self):
        """
        Test that /static/js/vendor/* only serves .js files.
        
        Security: Vendor JS route should not serve non-JS files.
        Expects: 400 or 404 for .json and .txt files in vendor directory.
        """
        response = self.client.get('/static/js/vendor/config.json')
        self.assertIn(response.status_code, [400, 404])
        
        response = self.client.get('/static/js/vendor/secrets.txt')
        self.assertIn(response.status_code, [400, 404])


class TestAcronymChecker(unittest.TestCase):
    """Test acronym checker edge cases."""
    
    def setUp(self):
        """Set up test fixtures."""
        pass
    
    def test_acronym_in_parentheses_is_counted(self):
        """
        Test that acronyms in parentheses are counted as usage.
        
        Functionality: (FMEA) format is valid acronym usage.
        Expects: FMEA found in usage dict with count >= 1.
        
        v3.0.9: Fixed bug where (FMEA) was skipped entirely.
        v3.0.10: Fixed test to use correct method name and argument format.
        """
        try:
            from acronym_checker import AcronymChecker
        except ImportError:
            self.skipTest("acronym_checker module not available")
        
        checker = AcronymChecker()
        
        # _find_usage expects List[Tuple[int, str]] - (paragraph_index, text)
        test_paragraphs = [(0, "(FMEA) is required for this procedure.")]
        
        # Call correct internal method with correct format
        usage = checker._find_usage(test_paragraphs)
        
        # FMEA should be found (not skipped due to parentheses)
        self.assertIn('FMEA', usage, 
            "Acronym in parentheses should be counted as usage")
        self.assertEqual(usage['FMEA'].usage_count, 1)
    
    def test_inline_definition_marks_as_defined(self):
        """
        Test that 'Full Name (ACRONYM)' pattern marks acronym as defined.
        
        Functionality: Inline definitions should mark acronyms as defined.
        Expects: FMEA in checker._defined set after extraction.
        """
        try:
            from acronym_checker import AcronymChecker
        except ImportError:
            self.skipTest("acronym_checker module not available")
        
        checker = AcronymChecker()
        
        # Extract inline definitions
        test_text = "Failure Mode and Effects Analysis (FMEA) is a technique."
        checker._extract_inline_definitions(test_text)
        
        # FMEA should now be in the defined set
        self.assertIn('FMEA', checker._defined,
            "Inline definition pattern should mark acronym as defined")
    
    def test_doc_refs_are_skipped(self):
        """
        Test that document references like ACR-123 are not flagged.
        
        Functionality: Doc ref patterns (XXX-###) should be ignored.
        Expects: ACR not in usage when followed by hyphen-number.
        """
        try:
            from acronym_checker import AcronymChecker
        except ImportError:
            self.skipTest("acronym_checker module not available")
        
        checker = AcronymChecker()
        
        # _find_usage expects List[Tuple[int, str]] - (paragraph_index, text)
        test_paragraphs = [(0, "See document ACR-123 for details.")]
        usage = checker._find_usage(test_paragraphs)
        
        # ACR should NOT be in usage (it's a doc ref pattern)
        self.assertNotIn('ACR', usage,
            "Document references (ACR-123) should be skipped")
    
    def test_strict_mode_flags_common_acronyms(self):
        """
        Test that strict mode (ignore_common_acronyms=False) flags NASA.
        
        Functionality: Strict mode requires all acronyms to be defined.
        Expects: NASA flagged as undefined in strict mode.
        
        v3.0.33: Added for strict mode testing.
        """
        try:
            from acronym_checker import AcronymChecker
        except ImportError:
            self.skipTest("acronym_checker module not available")
        
        # Create checker in strict mode (ignore_common_acronyms=False)
        checker = AcronymChecker(ignore_common_acronyms=False)
        
        # Test paragraph with NASA (a common acronym in UNIVERSAL_SKIP)
        test_paragraphs = [(0, "NASA launched a new satellite.")]
        
        # Run the check
        issues = checker.check(test_paragraphs, full_text="NASA launched a new satellite.")
        
        # In strict mode, NASA should be flagged because it's not defined in document
        self.assertEqual(len(issues), 1, 
            "Strict mode should flag NASA as undefined")
        self.assertIn('NASA', issues[0]['message'],
            "NASA should be in the issue message")
        
        # Verify metrics
        metrics = checker.get_metrics()
        self.assertEqual(metrics['flagged_count'], 1)
        self.assertEqual(metrics['suppressed_by_allowlist_count'], 0)
        self.assertTrue(metrics['strict_mode'])
    
    def test_permissive_mode_suppresses_common_acronyms(self):
        """
        Test that permissive mode (ignore_common_acronyms=True) suppresses NASA.
        
        Functionality: Permissive mode allows common acronyms without definition.
        Expects: NASA not flagged, suppressed_by_allowlist_count = 1.
        
        v3.0.33: Added for permissive mode testing.
        """
        try:
            from acronym_checker import AcronymChecker
        except ImportError:
            self.skipTest("acronym_checker module not available")
        
        # Create checker in permissive mode (ignore_common_acronyms=True)
        checker = AcronymChecker(ignore_common_acronyms=True)
        
        # Test paragraph with NASA (a common acronym in UNIVERSAL_SKIP)
        test_paragraphs = [(0, "NASA launched a new satellite.")]
        
        # Run the check
        issues = checker.check(test_paragraphs, full_text="NASA launched a new satellite.")
        
        # In permissive mode, NASA should NOT be flagged
        self.assertEqual(len(issues), 0, 
            "Permissive mode should suppress NASA")
        
        # Verify metrics
        metrics = checker.get_metrics()
        self.assertEqual(metrics['flagged_count'], 0)
        self.assertEqual(metrics['suppressed_by_allowlist_count'], 1)
        self.assertIn('NASA', metrics['allowlist_matches'])
        self.assertFalse(metrics['strict_mode'])
    
    def test_metrics_track_defined_acronyms(self):
        """
        Test that metrics correctly track defined acronyms.
        
        Functionality: Metrics should count defined vs flagged acronyms.
        Expects: defined_count=1, flagged_count=0 for text with inline def.
        
        v3.0.33: Added for metrics testing.
        """
        try:
            from acronym_checker import AcronymChecker
        except ImportError:
            self.skipTest("acronym_checker module not available")
        
        checker = AcronymChecker(ignore_common_acronyms=False)
        
        # Test text with an inline definition + another usage
        test_text = "Failure Mode and Effects Analysis (FMEA) is used. FMEA is important."
        test_paragraphs = [(0, test_text)]
        
        # Run check
        issues = checker.check(test_paragraphs, full_text=test_text)
        
        # FMEA is defined inline, so no issues
        self.assertEqual(len(issues), 0, 
            "Defined acronyms should not be flagged")
        
        # Verify metrics
        metrics = checker.get_metrics()
        self.assertEqual(metrics['defined_count'], 1)
        self.assertEqual(metrics['flagged_count'], 0)


class TestRoleDeliverableSeparation(unittest.TestCase):
    """
    Tests for role vs deliverable classification and export.
    
    v3.0.12: These tests ensure deliverables never appear in role exports.
    """
    
    def test_role_classification(self):
        """
        Test that roles are classified correctly.
        
        Functionality: Role extractor should identify personnel roles.
        Expects: Terms like 'Program Manager', 'Test Engineer' classified as role.
        """
        try:
            from role_extractor_v3 import RoleExtractor
        except ImportError:
            self.skipTest("role_extractor_v3 module not available")
        
        extractor = RoleExtractor()
        
        # Clear roles should be classified as 'role'
        roles = ['Program Manager', 'Test Engineer', 'Quality Assurance Lead', 'Systems Engineer']
        for role in roles:
            result = extractor.classify_extraction(role)
            self.assertEqual(result['type'], 'role', 
                f"'{role}' should be classified as role, got '{result['type']}'")
    
    def test_deliverable_classification(self):
        """
        Test that deliverables are classified correctly.
        
        Functionality: Role extractor should identify document deliverables.
        Expects: Terms like 'Test Plan', 'System Specification' classified as deliverable.
        """
        try:
            from role_extractor_v3 import RoleExtractor
        except ImportError:
            self.skipTest("role_extractor_v3 module not available")
        
        extractor = RoleExtractor()
        
        # Deliverables should be classified as 'deliverable'
        deliverables = ['Test Plan', 'System Specification', 'Verification Matrix', 
                       'Requirements Document', 'Design Report']
        for deliv in deliverables:
            result = extractor.classify_extraction(deliv)
            self.assertEqual(result['type'], 'deliverable', 
                f"'{deliv}' should be classified as deliverable, got '{result['type']}'")
    
    def test_role_suffix_wins_tiebreak(self):
        """
        Test that role suffix wins when both indicators present.
        
        Functionality: 'Manager', 'Engineer' suffixes indicate roles.
        Expects: 'Test Manager', 'Report Engineer' classified as role.
        """
        try:
            from role_extractor_v3 import RoleExtractor
        except ImportError:
            self.skipTest("role_extractor_v3 module not available")
        
        extractor = RoleExtractor()
        
        # Role suffix should win over deliverable-like prefix
        result = extractor.classify_extraction('Test Manager')
        self.assertEqual(result['type'], 'role', 
            "'Test Manager' should be role (Manager suffix wins)")
        
        result = extractor.classify_extraction('Report Engineer')
        self.assertEqual(result['type'], 'role', 
            "'Report Engineer' should be role (Engineer suffix wins)")
    
    def test_entity_kind_populated(self):
        """
        Test that ExtractedRole has entity_kind populated.
        
        Functionality: Role objects must have entity_kind for filtering.
        Expects: entity_kind is EntityKind enum for all extracted entities.
        """
        try:
            from role_extractor_v3 import RoleExtractor, EntityKind
        except ImportError:
            self.skipTest("role_extractor_v3 module not available")
        
        extractor = RoleExtractor()
        
        # Extract from sample text with both roles and deliverables
        text = "The Program Manager shall review the Test Plan. The Test Engineer is responsible."
        extracted = extractor.extract_from_text(text, "test")
        
        # Check that entity_kind is populated
        for name, role in extracted.items():
            self.assertTrue(hasattr(role, 'entity_kind'), 
                f"'{name}' should have entity_kind attribute")
            self.assertIsInstance(role.entity_kind, EntityKind,
                f"'{name}' entity_kind should be EntityKind enum")
    
    def test_export_filter_simulation(self):
        """
        Test that role export filtering works correctly.
        
        Functionality: Export filter separates roles from deliverables.
        Expects: Role export contains only role entities, not deliverables.
        """
        # Simulate the export filter logic
        mock_entities = {
            'roles': [
                {'canonical_name': 'Program Manager', 'entity_kind': 'role'},
                {'canonical_name': 'Test Engineer', 'entity_kind': 'role'}
            ],
            'deliverables': [
                {'canonical_name': 'Test Plan', 'entity_kind': 'deliverable'},
                {'canonical_name': 'System Specification', 'entity_kind': 'deliverable'}
            ]
        }
        
        # Role export should only include roles
        role_export = mock_entities['roles']
        self.assertEqual(len(role_export), 2)
        
        role_names = [r['canonical_name'] for r in role_export]
        self.assertIn('Program Manager', role_names)
        self.assertIn('Test Engineer', role_names)
        self.assertNotIn('Test Plan', role_names)
        self.assertNotIn('System Specification', role_names)
    
    def test_deliverables_excluded_from_roles(self):
        """
        Critical test: Deliverables must never appear in role exports.
        
        Security: Prevents data contamination in role-specific exports.
        Expects: After filtering, only roles remain (no deliverables).
        """
        # This simulates the legacy fallback filter
        mock_roles = {
            'program manager': {'canonical_name': 'Program Manager', 'entity_kind': 'role'},
            'test plan': {'canonical_name': 'Test Plan', 'entity_kind': 'deliverable'},
            'systems engineer': {'canonical_name': 'Systems Engineer', 'entity_kind': 'role'},
            'verification matrix': {'canonical_name': 'Verification Matrix', 'entity_kind': 'deliverable'}
        }
        
        # Filter as export would
        filtered = {k: v for k, v in mock_roles.items() 
                   if v.get('entity_kind') != 'deliverable'}
        
        self.assertEqual(len(filtered), 2, "Should have 2 roles after filtering")
        self.assertIn('program manager', filtered)
        self.assertIn('systems engineer', filtered)
        self.assertNotIn('test plan', filtered, "Test Plan should be excluded")
        self.assertNotIn('verification matrix', filtered, "Verification Matrix should be excluded")
    
    def test_unknown_excluded_from_export(self):
        """
        v3.0.12b: Unknown entities must be excluded from role export.
        
        Functionality: Only explicit 'role' kind should be exported.
        Expects: Unknown entity_kind filtered out of role exports.
        """
        mock_entities = {
            'role1': {'canonical_name': 'Program Manager', 'entity_kind': 'role'},
            'deliv1': {'canonical_name': 'Test Plan', 'entity_kind': 'deliverable'},
            'unknown1': {'canonical_name': 'Ambiguous Term', 'entity_kind': 'unknown'},
            'role2': {'canonical_name': 'Test Engineer', 'entity_kind': 'role'}
        }
        
        # Export filter: only explicit 'role' kind
        exported = {k: v for k, v in mock_entities.items() 
                   if v.get('entity_kind') == 'role'}
        
        self.assertEqual(len(exported), 2, "Should have 2 roles")
        self.assertIn('role1', exported)
        self.assertIn('role2', exported)
        self.assertNotIn('unknown1', exported, "Unknown should be excluded from export")
        self.assertNotIn('deliv1', exported, "Deliverable should be excluded")
    
    def test_empty_export_when_only_deliverables(self):
        """
        If document has only deliverables, role export should be empty.
        
        Edge Case: Documents with no roles should produce empty role exports.
        Expects: roles list is empty when only deliverables present.
        """
        mock_entities = {
            'roles': [],
            'deliverables': [
                {'canonical_name': 'Test Plan', 'entity_kind': 'deliverable'},
                {'canonical_name': 'System Spec', 'entity_kind': 'deliverable'}
            ],
            'unknown': []
        }
        
        # Roles export should be empty
        self.assertEqual(len(mock_entities['roles']), 0)
    
    def test_classification_persistence(self):
        """
        Classification must be persisted at creation, not inferred at export.
        
        Functionality: entity_kind set during extraction, not on-demand.
        Expects: entity_kind, kind_confidence, and kind_reason all set.
        """
        try:
            from role_extractor_v3 import RoleExtractor, EntityKind
        except ImportError:
            self.skipTest("role_extractor_v3 module not available")
        
        extractor = RoleExtractor()
        
        # Create a role entry through the helper
        role = extractor._create_extracted_role("Test Plan", "Test Plan")
        
        # Classification should be persisted
        self.assertEqual(role.entity_kind, EntityKind.DELIVERABLE,
            "Classification should be persisted at creation")
        self.assertGreater(role.kind_confidence, 0,
            "Confidence should be set at creation")
        self.assertTrue(len(role.kind_reason) > 0,
            "Reason should be set at creation")
    
    def test_edge_case_lead_test_plan(self):
        """
        Edge case: 'Lead Test Plan' should be deliverable (Plan suffix wins).
        
        Classification Logic: Document suffix takes precedence over 'Lead'.
        Expects: 'Lead Test Plan' classified as deliverable.
        """
        try:
            from role_extractor_v3 import RoleExtractor
        except ImportError:
            self.skipTest("role_extractor_v3 module not available")
        
        extractor = RoleExtractor()
        
        # 'Lead Test Plan' - 'Plan' suffix should win
        result = extractor.classify_extraction('Lead Test Plan')
        self.assertEqual(result['type'], 'deliverable',
            "'Lead Test Plan' should be deliverable (Plan suffix wins)")


class TestHyperlinkHealth(unittest.TestCase):
    """
    Tests for hyperlink health validation.
    
    v3.0.33 Chunk B: Added for hyperlink health feature testing.
    """
    
    def test_offline_mode_returns_valid_format(self):
        """
        Test that offline mode returns VALID_FORMAT (not VALID) for web URLs.
        
        Functionality: Offline mode validates format only, not reachability.
        Expects: status = VALID_FORMAT, message contains 'offline'.
        
        v3.0.33 Chunk B: Offline mode should not claim URLs are VALID,
        only that their format is valid.
        """
        try:
            from hyperlink_health import HyperlinkHealthValidator, HealthMode, LinkStatus
        except ImportError:
            self.skipTest("hyperlink_health module not available")
        
        validator = HyperlinkHealthValidator(mode=HealthMode.OFFLINE)
        
        # Validate a web URL in offline mode
        record = validator.validate_link(
            target="https://example.com/test",
            display_text="Example Link"
        )
        
        # Should be VALID_FORMAT, not VALID (we didn't actually check the URL)
        self.assertEqual(record.status, LinkStatus.VALID_FORMAT.value,
            "Offline mode should return VALID_FORMAT for web URLs, not VALID")
        self.assertIn("offline", record.status_message.lower())
    
    def test_offline_mode_report_shape(self):
        """
        Test that offline mode report has correct shape.
        
        API Design: Report structure must match frontend expectations.
        Expects: Report contains summary with total_links, links array, validation_mode.
        """
        try:
            from hyperlink_health import validate_document_links, HyperlinkHealthReport
        except ImportError:
            self.skipTest("hyperlink_health module not available")
        
        # Create a mock DOCX for testing would be complex, 
        # so we test the report generation directly
        try:
            from hyperlink_health import HyperlinkHealthValidator, HealthMode
        except ImportError:
            self.skipTest("hyperlink_health module not available")
        
        validator = HyperlinkHealthValidator(mode=HealthMode.OFFLINE)
        
        # Validate some test links
        test_links = [
            {'target': 'https://example.com', 'display_text': 'Example'},
            {'target': 'mailto:test@example.com', 'display_text': 'Email'},
            {'target': '#bookmark1', 'display_text': 'Bookmark'},
        ]
        
        validator.validate_batch(test_links)
        report = validator.generate_report(document_path='/test/doc.docx', document_name='doc.docx')
        
        # Check report shape
        report_dict = report.to_dict()
        self.assertIn('summary', report_dict)
        self.assertIn('total_links', report_dict['summary'])
        self.assertIn('links', report_dict)
        self.assertEqual(report_dict['summary']['total_links'], 3)
        self.assertEqual(report_dict['validation_mode'], 'offline')
    
    def test_link_type_classification(self):
        """
        Test that link types are correctly classified.
        
        Functionality: Different URL formats are categorized correctly.
        Expects: Web URLs, bookmarks, mailto, file paths all classified properly.
        """
        try:
            from hyperlink_health import HyperlinkHealthValidator, HealthMode, LinkType
        except ImportError:
            self.skipTest("hyperlink_health module not available")
        
        validator = HyperlinkHealthValidator(mode=HealthMode.OFFLINE)
        
        # Test different link types
        test_cases = [
            ('https://example.com', LinkType.WEB_URL),
            ('http://test.org/page', LinkType.WEB_URL),
            ('#bookmark', LinkType.INTERNAL_BOOKMARK),
            ('mailto:test@example.com', LinkType.MAILTO),
            ('C:\\Documents\\file.pdf', LinkType.FILE_PATH),
            ('\\\\server\\share\\file.docx', LinkType.NETWORK_PATH),
            ('', LinkType.EMPTY),
        ]
        
        for target, expected_type in test_cases:
            result = validator.classify_link(target)
            self.assertEqual(result, expected_type,
                f"'{target}' should be classified as {expected_type.value}")
    
    def test_ps1_validator_mode_available(self):
        """
        Test that PS1_VALIDATOR mode is recognized.
        
        Functionality: PowerShell validator mode for air-gapped environments.
        Expects: HealthMode.PS1_VALIDATOR exists with value 'ps1_validator'.
        """
        try:
            from hyperlink_health import HealthMode
        except ImportError:
            self.skipTest("hyperlink_health module not available")
        
        # Verify PS1_VALIDATOR mode exists
        self.assertTrue(hasattr(HealthMode, 'PS1_VALIDATOR'))
        self.assertEqual(HealthMode.PS1_VALIDATOR.value, 'ps1_validator')
    
    def test_link_status_enum_completeness(self):
        """
        Test that all required link statuses are available.
        
        API Design: LinkStatus enum must cover all possible validation results.
        Expects: All required status values exist in LinkStatus enum.
        """
        try:
            from hyperlink_health import LinkStatus
        except ImportError:
            self.skipTest("hyperlink_health module not available")
        
        # v3.0.33: Check new statuses added for PS1 validator
        required_statuses = [
            'WORKING', 'BROKEN', 'BLOCKED', 'UNKNOWN', 'VALID_FORMAT',
            'VALID', 'INVALID', 'WARNING', 'TIMEOUT', 'NOT_FOUND'
        ]
        
        for status in required_statuses:
            self.assertTrue(hasattr(LinkStatus, status),
                f"LinkStatus should have {status} enum value")
    
    def test_ps1_validator_script_exists(self):
        """
        Test that HyperlinkValidator.ps1 script exists in tools folder.
        
        Deployment: PS1 script must be shipped with application.
        Expects: Script found by _find_ps1_validator() and file exists.
        
        v3.0.37: PS1 validator script must be shipped with the application.
        """
        try:
            from hyperlink_health import HyperlinkHealthValidator, HealthMode
        except ImportError:
            self.skipTest("hyperlink_health module not available")
        
        validator = HyperlinkHealthValidator(mode=HealthMode.OFFLINE)
        ps1_path = validator._find_ps1_validator()
        
        self.assertIsNotNone(ps1_path, 
            "HyperlinkValidator.ps1 script should be found in tools folder")
        
        import os
        self.assertTrue(os.path.exists(ps1_path),
            f"PS1 script path should exist: {ps1_path}")
    
    def test_ps1_validator_script_has_required_content(self):
        """
        Test that HyperlinkValidator.ps1 has correct structure.
        
        Contract: Script must define expected parameters and output format.
        Expects: InputFile, OutputFormat, TimeoutSeconds params and status values.
        
        v3.0.37: Script must define expected parameters and output format.
        """
        try:
            from hyperlink_health import HyperlinkHealthValidator, HealthMode
        except ImportError:
            self.skipTest("hyperlink_health module not available")
        
        validator = HyperlinkHealthValidator(mode=HealthMode.OFFLINE)
        ps1_path = validator._find_ps1_validator()
        
        if not ps1_path:
            self.skipTest("PS1 validator not available")
        
        with open(ps1_path, 'r') as f:
            content = f.read()
        
        # Check for required parameters
        self.assertIn('$InputFile', content, 
            "Script should have InputFile parameter")
        self.assertIn('$OutputFormat', content, 
            "Script should have OutputFormat parameter")
        self.assertIn('$TimeoutSeconds', content, 
            "Script should have TimeoutSeconds parameter")
        
        # Check for status values
        required_statuses = ['WORKING', 'BROKEN', 'TIMEOUT', 'BLOCKED', 'UNKNOWN']
        for status in required_statuses:
            self.assertIn(status, content,
                f"Script should define {status} status")


class TestHyperlinkConfigEndpoint(unittest.TestCase):
    """
    Tests for hyperlink config API endpoint.
    
    v3.0.37: Tests for /api/config/hyperlinks endpoint.
    """
    
    def setUp(self):
        """Set up test client."""
        try:
            from app import app
            self.app = app
            self.client = app.test_client()
            app.config['TESTING'] = True
        except ImportError:
            self.skipTest("app module not available")
    
    def test_get_hyperlink_config(self):
        """
        Test GET /api/config/hyperlinks returns settings.
        
        Functionality: UI needs hyperlink validation configuration.
        Expects: 200 response with validation_mode, modes, ps1_available.
        """
        response = self.client.get('/api/config/hyperlinks')
        self.assertEqual(response.status_code, 200)
        
        data = response.get_json()
        self.assertTrue(data.get('success'))
        self.assertIn('validation_mode', data.get('data', {}))
        self.assertIn('modes', data.get('data', {}))
        self.assertIn('ps1_available', data.get('data', {}))


class TestCommentInserter(unittest.TestCase):
    """
    Tests for comment inserter module.
    
    v3.0.37 Batch G: Tests for hyperlink comment insertion.
    """
    
    def test_hyperlink_comment_creation(self):
        """
        Test HyperlinkComment dataclass creates proper comment text.
        
        Functionality: Comment text should include status and URL.
        Expects: Comment text contains 'BROKEN' and the target URL.
        """
        try:
            from comment_inserter import HyperlinkComment
        except ImportError:
            self.skipTest("comment_inserter module not available")
        
        comment = HyperlinkComment(
            target_url='https://example.com/broken',
            display_text='Test Link',
            status='broken',
            status_message='HTTP 404 Not Found'
        )
        
        self.assertIn('BROKEN', comment.comment_text)
        self.assertIn('https://example.com/broken', comment.comment_text)
    
    def test_hyperlink_comment_status_mapping(self):
        """
        Test that different statuses generate appropriate comments.
        
        Functionality: Status determines comment label text.
        Expects: Each status maps to expected text in comment.
        """
        try:
            from comment_inserter import HyperlinkComment
        except ImportError:
            self.skipTest("comment_inserter module not available")
        
        test_cases = [
            ('broken', 'BROKEN LINK'),
            ('timeout', 'TIMEOUT'),
            ('blocked', 'BLOCKED'),
            ('dns_failed', 'DNS FAILED'),
            ('ssl_error', 'SSL ERROR'),
        ]
        
        for status, expected_text in test_cases:
            comment = HyperlinkComment(
                target_url='https://test.com',
                display_text='Test',
                status=status,
                status_message='Test message'
            )
            self.assertIn(expected_text, comment.comment_text,
                f"Status '{status}' should produce comment containing '{expected_text}'")
    
    def test_generate_comment_pack(self):
        """
        Test comment pack generation.
        
        Functionality: Comment pack bundles all broken link comments.
        Expects: Pack contains header, document name, all broken link URLs.
        """
        try:
            from comment_inserter import generate_comment_pack
        except ImportError:
            self.skipTest("comment_inserter module not available")
        
        broken_links = [
            {'target': 'https://broken1.com', 'display_text': 'Link 1', 
             'status': 'broken', 'status_message': 'HTTP 404'},
            {'target': 'https://timeout.com', 'display_text': 'Link 2', 
             'status': 'timeout', 'status_message': 'Timed out'},
        ]
        
        pack = generate_comment_pack(broken_links, 'TestDoc.docx')
        
        self.assertIn('HYPERLINK COMMENT PACK', pack)
        self.assertIn('TestDoc.docx', pack)
        self.assertIn('https://broken1.com', pack)
        self.assertIn('https://timeout.com', pack)
        self.assertIn('BROKEN LINK', pack)
        self.assertIn('TIMEOUT', pack)
    
    def test_comment_pack_skips_valid_links(self):
        """
        Test that comment pack skips links with OK status.
        
        Functionality: Only non-working links should be commented.
        Expects: Working links excluded from comment pack.
        """
        try:
            from comment_inserter import generate_comment_pack
        except ImportError:
            self.skipTest("comment_inserter module not available")
        
        links = [
            {'target': 'https://working.com', 'display_text': 'Good Link', 
             'status': 'working', 'status_message': 'OK'},
            {'target': 'https://broken.com', 'display_text': 'Bad Link', 
             'status': 'broken', 'status_message': 'HTTP 404'},
        ]
        
        pack = generate_comment_pack(links, 'Test.docx')
        
        # Should include broken link
        self.assertIn('https://broken.com', pack)
        # Should NOT include working link in the comment entries
        # (it may appear in summary stats but not as a comment suggestion)
        self.assertNotIn('Good Link', pack.split('SUGGESTED COMMENT')[0] if 'SUGGESTED COMMENT' in pack else pack)
    
    def test_comment_scope_broken_only(self):
        """
        Test CommentScope.BROKEN_ONLY filters correctly.
        
        Functionality: BROKEN_ONLY scope includes only definitively broken links.
        Expects: broken/404/invalid included, timeout/blocked/unknown excluded.
        
        v3.0.39: Added for scope filtering validation.
        """
        try:
            from comment_inserter import should_comment_link, CommentScope
        except ImportError:
            self.skipTest("comment_inserter module not available")
        
        # BROKEN_ONLY should include broken, 404, invalid, dns_failed, ssl_error
        self.assertTrue(should_comment_link('broken', CommentScope.BROKEN_ONLY))
        self.assertTrue(should_comment_link('404', CommentScope.BROKEN_ONLY))
        self.assertTrue(should_comment_link('invalid', CommentScope.BROKEN_ONLY))
        self.assertTrue(should_comment_link('dns_failed', CommentScope.BROKEN_ONLY))
        self.assertTrue(should_comment_link('ssl_error', CommentScope.BROKEN_ONLY))
        
        # BROKEN_ONLY should exclude timeout, blocked, unknown
        self.assertFalse(should_comment_link('timeout', CommentScope.BROKEN_ONLY))
        self.assertFalse(should_comment_link('blocked', CommentScope.BROKEN_ONLY))
        self.assertFalse(should_comment_link('unknown', CommentScope.BROKEN_ONLY))
        
        # Both scopes should exclude valid/working
        self.assertFalse(should_comment_link('valid', CommentScope.BROKEN_ONLY))
        self.assertFalse(should_comment_link('working', CommentScope.BROKEN_ONLY))
    
    def test_comment_scope_all_non_working(self):
        """
        Test CommentScope.ALL_NON_WORKING includes timeout/blocked/unknown.
        
        Functionality: ALL_NON_WORKING scope includes any non-OK status.
        Expects: All non-working statuses included, only valid/working excluded.
        
        v3.0.39: Added for scope filtering validation.
        """
        try:
            from comment_inserter import should_comment_link, CommentScope
        except ImportError:
            self.skipTest("comment_inserter module not available")
        
        # ALL_NON_WORKING should include everything non-OK
        self.assertTrue(should_comment_link('broken', CommentScope.ALL_NON_WORKING))
        self.assertTrue(should_comment_link('timeout', CommentScope.ALL_NON_WORKING))
        self.assertTrue(should_comment_link('blocked', CommentScope.ALL_NON_WORKING))
        self.assertTrue(should_comment_link('unknown', CommentScope.ALL_NON_WORKING))
        
        # Still excludes valid/working
        self.assertFalse(should_comment_link('valid', CommentScope.ALL_NON_WORKING))
        self.assertFalse(should_comment_link('working', CommentScope.ALL_NON_WORKING))
        self.assertFalse(should_comment_link('ok', CommentScope.ALL_NON_WORKING))
    
    def test_hyperlink_location_describe(self):
        """
        Test HyperlinkLocation.describe() produces useful output.
        
        Functionality: Location hints help users find links in document.
        Expects: describe() returns human-readable location text.
        
        v3.0.39: Added for location hints validation.
        """
        try:
            from comment_inserter import HyperlinkLocation
        except ImportError:
            self.skipTest("comment_inserter module not available")
        
        # Test paragraph location
        para_loc = HyperlinkLocation(location_type='paragraph', paragraph_index=5)
        self.assertIn('Paragraph 6', para_loc.describe())  # 1-indexed
        
        # Test table location
        table_loc = HyperlinkLocation(
            location_type='table', table_index=0, row_index=2, cell_index=1
        )
        desc = table_loc.describe()
        self.assertIn('Table 1', desc)
        self.assertIn('Row 3', desc)
        self.assertIn('Cell 2', desc)
        
        # Test header location
        header_loc = HyperlinkLocation(location_type='header', section_index=0)
        self.assertIn('Header', header_loc.describe())
        self.assertIn('Section 1', header_loc.describe())
        
        # Test footer location
        footer_loc = HyperlinkLocation(location_type='footer', section_index=1)
        self.assertIn('Footer', footer_loc.describe())
        self.assertIn('Section 2', footer_loc.describe())
    
    def test_comment_pack_includes_location_hints(self):
        """
        Test comment pack includes location hints when hyperlink_info provided.
        
        Functionality: Location context helps users navigate to broken links.
        Expects: Pack includes Table, Row, Cell info and surrounding context.
        
        v3.0.39: Added for location hints validation.
        """
        try:
            from comment_inserter import generate_comment_pack, HyperlinkLocation
        except ImportError:
            self.skipTest("comment_inserter module not available")
        
        broken_links = [
            {
                'target': 'https://broken.com',
                'display_text': 'Broken Link',
                'status': 'broken',
                'status_message': 'HTTP 404',
                'hyperlink_info': {
                    'location': HyperlinkLocation(
                        location_type='table',
                        table_index=2,
                        row_index=5,
                        cell_index=3
                    ),
                    'surrounding_text': 'context around'
                }
            }
        ]
        
        pack = generate_comment_pack(broken_links, 'Test.docx')
        
        # Should include location info
        self.assertIn('Table 3', pack)  # 1-indexed
        self.assertIn('Row 6', pack)
        self.assertIn('Cell 4', pack)
        self.assertIn('CONTEXT:', pack)
        self.assertIn('context around', pack)
    
    def test_docx_comment_insertion_integration(self):
        """
        Integration test: create DOCX with hyperlink, insert comments, verify XML.
        
        Integration: Full workflow from document to commented output.
        Expects: Output DOCX exists, has comments, contains comment markers in XML.
        
        v3.0.39: Added integration test for actual DOCX comment insertion.
        This test:
        1. Creates a temp DOCX with a hyperlink
        2. Runs insert_comments_at_hyperlinks
        3. Verifies the output DOCX contains comment markers in document.xml
        """
        try:
            from docx import Document
            from docx.oxml.ns import qn
            from comment_inserter import insert_comments_at_hyperlinks, CommentScope
            import zipfile
        except ImportError as e:
            self.skipTest(f"Required module not available: {e}")
        
        import tempfile
        import os
        
        # Create a temp DOCX with a hyperlink
        with tempfile.TemporaryDirectory() as tmpdir:
            # Create source document
            source_path = os.path.join(tmpdir, 'test_source.docx')
            doc = Document()
            para = doc.add_paragraph('Click here: ')
            
            # Add hyperlink using python-docx's low-level API
            from docx.oxml import OxmlElement
            from docx.opc.constants import RELATIONSHIP_TYPE as RT
            
            # Add external relationship for the hyperlink
            r_id = doc.part.relate_to(
                'https://broken-test-link.example.com/404',
                RT.HYPERLINK,
                is_external=True
            )
            
            # Create hyperlink element
            hyperlink = OxmlElement('w:hyperlink')
            hyperlink.set(qn('r:id'), r_id)
            
            # Add run with text inside hyperlink
            run_elem = OxmlElement('w:r')
            text_elem = OxmlElement('w:t')
            text_elem.text = 'Test Hyperlink'
            run_elem.append(text_elem)
            hyperlink.append(run_elem)
            
            # Append hyperlink to paragraph
            para._element.append(hyperlink)
            
            doc.save(source_path)
            
            # Define broken links to comment
            broken_links = [
                {
                    'target': 'https://broken-test-link.example.com/404',
                    'display_text': 'Test Hyperlink',
                    'status': 'broken',
                    'status_message': 'HTTP 404 Not Found'
                }
            ]
            
            # Run comment insertion
            output_path = os.path.join(tmpdir, 'test_commented.docx')
            result_path, count = insert_comments_at_hyperlinks(
                filepath=source_path,
                broken_links=broken_links,
                author='TestAuthor',
                output_path=output_path,
                scope=CommentScope.ALL_NON_WORKING
            )
            
            # Verify output file exists
            self.assertTrue(os.path.exists(result_path), "Output DOCX should exist")
            
            # Verify at least one comment was inserted
            self.assertGreaterEqual(count, 1, "Should have inserted at least 1 comment")
            
            # Open the DOCX as a zip and check for comment markers in document.xml
            with zipfile.ZipFile(result_path, 'r') as zf:
                # Read document.xml
                doc_xml = zf.read('word/document.xml').decode('utf-8')
                
                # Check for comment range markers (commentRangeStart, commentRangeEnd, commentReference)
                self.assertIn('commentRangeStart', doc_xml, 
                    "document.xml should contain commentRangeStart marker")
                self.assertIn('commentRangeEnd', doc_xml,
                    "document.xml should contain commentRangeEnd marker")
                self.assertIn('commentReference', doc_xml,
                    "document.xml should contain commentReference marker")


class TestStatementForgePersistence(unittest.TestCase):
    """
    Tests for Statement Forge persistence and auto-availability.
    
    v3.0.33 Chunk C: Added for SF session persistence.
    """
    
    def test_sf_routes_have_availability_endpoint(self):
        """
        Test that the availability endpoint function exists.
        
        Functionality: SF availability status needed by frontend.
        Expects: check_availability function exists and is callable.
        """
        try:
            from statement_forge.routes import check_availability
        except ImportError:
            self.skipTest("statement_forge.routes not available or check_availability not defined")
        
        # Verify the check_availability function exists and is callable
        self.assertTrue(callable(check_availability),
            "check_availability should be a callable function")
        
        # Check function is decorated with route
        # Flask routes have __name__ attribute
        self.assertEqual(check_availability.__name__, 'check_availability',
            "Function should be named check_availability")
    
    def test_sf_store_and_retrieve_statements(self):
        """
        Test that statements can be stored and retrieved from session.
        
        Functionality: SF persistence allows viewing previous extractions.
        Expects: _store_statements and _get_statements have correct signatures.
        """
        try:
            from statement_forge.routes import _store_statements, _get_statements, _session_statements
            from statement_forge.models import Statement
        except ImportError:
            self.skipTest("statement_forge modules not available")
        
        # Create test statements
        test_statements = [
            Statement(id='test1', title='Test 1', description='Description 1'),
            Statement(id='test2', title='Test 2', description='Description 2'),
        ]
        
        # This would require Flask app context in real tests
        # For now, just verify the functions exist and have correct signatures
        import inspect
        store_sig = inspect.signature(_store_statements)
        get_sig = inspect.signature(_get_statements)
        
        self.assertEqual(len(store_sig.parameters), 1, 
            "_store_statements should take 1 parameter (statements)")
        self.assertEqual(len(get_sig.parameters), 0, 
            "_get_statements should take no parameters")
    
    def test_statement_model_serialization(self):
        """
        Test Statement model can be serialized and deserialized.
        
        Functionality: Statements must serialize for session storage.
        Expects: to_dict() produces dict, from_dict() restores Statement.
        """
        try:
            from statement_forge.models import Statement
        except ImportError:
            self.skipTest("statement_forge.models not available")
        
        # Create a statement with all fields
        stmt = Statement(
            id='test123',
            number='4.1.1',
            title='Test Statement',
            description='The system shall do something.',
            level=2,
            section='4.1',
            directive='shall',
            role='System Engineer',
            notes=['Note 1', 'Note 2']
        )
        
        # Serialize
        data = stmt.to_dict()
        
        # Check required fields
        self.assertEqual(data['id'], 'test123')
        self.assertEqual(data['directive'], 'shall')
        self.assertEqual(data['role'], 'System Engineer')
        self.assertIsInstance(data['notes'], list)
        
        # Deserialize
        restored = Statement.from_dict(data)
        self.assertEqual(restored.id, stmt.id)
        self.assertEqual(restored.directive, stmt.directive)
        self.assertEqual(restored.role, stmt.role)
    
    def test_sf_summary_includes_statements_ready_flag(self):
        """
        Test that review response includes statements_ready flag.
        
        API Design: Frontend needs to know if statements are pre-extracted.
        Expects: SF summary contains available, statements_ready, total_statements.
        
        v3.0.33 Chunk C: The review response should include a flag
        indicating whether statements were pre-extracted.
        """
        # This test validates the contract, actual integration
        # would require Flask app context
        expected_summary_fields = [
            'available',
            'statements_ready',
            'total_statements',
        ]
        
        # Mock summary as would be returned by review endpoint
        mock_summary = {
            'available': True,
            'statements_ready': True,
            'total_statements': 42,
            'directive_counts': {'shall': 20, 'must': 10, 'will': 12},
            'top_roles': ['System Engineer', 'Test Engineer'],
            'section_count': 5
        }
        
        for field in expected_summary_fields:
            self.assertIn(field, mock_summary,
                f"SF summary should include '{field}' field")


class TestExportEnhancements(unittest.TestCase):
    """
    Tests for v3.0.33 Chunk D: Export Enhancements.
    
    Validates:
    - Action Item column presence
    - Timestamped filename generation
    - Document metadata in export
    - Severity filtering
    """
    
    def test_timestamped_filename_generation(self):
        """
        Test that timestamped filenames are generated correctly.
        
        Functionality: Unique filenames prevent overwrites.
        Expects: Filename matches pattern: prefix_YYYYMMDD_HHMMSS.ext
        """
        from export_module import generate_timestamped_filename
        
        filename = generate_timestamped_filename('review_export', 'xlsx')
        
        # Should match pattern: review_export_YYYYMMDD_HHMMSS.xlsx
        import re
        pattern = r'^review_export_\d{8}_\d{6}\.xlsx$'
        self.assertRegex(filename, pattern,
            f"Filename '{filename}' should match timestamped pattern")
    
    def test_xlsx_export_has_action_item_column(self):
        """
        Test that XLSX export includes Action Item column.
        
        Functionality: Action Item column for tracking remediation.
        Expects: Issues sheet has 'Action Item' in header row.
        """
        from export_module import ExcelExporter
        
        # Mock results with issues
        results = {
            'score': 85,
            'grade': 'B',
            'issues': [
                {'severity': 'High', 'category': 'Grammar', 'message': 'Test issue 1'},
                {'severity': 'Medium', 'category': 'Style', 'message': 'Test issue 2'},
            ],
            'document_info': {'filename': 'test.docx', 'word_count': 100},
            'by_severity': {'High': 1, 'Medium': 1}
        }
        
        exporter = ExcelExporter()
        content = exporter.export(results)
        
        # Parse the workbook to verify Action Item column
        from openpyxl import load_workbook
        import io
        
        wb = load_workbook(io.BytesIO(content))
        
        # Check Issues sheet exists
        self.assertIn('Issues', wb.sheetnames, "Issues sheet should exist")
        
        ws = wb['Issues']
        
        # Check header row for Action Item column
        headers = [cell.value for cell in ws[1]]
        self.assertIn('Action Item', headers,
            f"Headers {headers} should include 'Action Item' column")
    
    def test_severity_filter_in_export(self):
        """
        Test that severity filter works in export.
        
        Functionality: Users can export only selected severity levels.
        Expects: Export excludes Low severity when filtering to Critical/High.
        """
        from export_module import ExcelExporter
        
        results = {
            'score': 85,
            'issues': [
                {'severity': 'Critical', 'category': 'Security', 'message': 'Critical issue'},
                {'severity': 'High', 'category': 'Grammar', 'message': 'High issue'},
                {'severity': 'Low', 'category': 'Style', 'message': 'Low issue'},
            ],
            'document_info': {'filename': 'test.docx'},
            'by_severity': {'Critical': 1, 'High': 1, 'Low': 1}
        }
        
        exporter = ExcelExporter()
        # Export only Critical and High
        content = exporter.export(results, severities=['Critical', 'High'])
        
        from openpyxl import load_workbook
        import io
        
        wb = load_workbook(io.BytesIO(content))
        ws = wb['Issues']
        
        # Count data rows (excluding header)
        data_rows = list(ws.iter_rows(min_row=2, max_col=3, values_only=True))
        severities_exported = [row[1] for row in data_rows if row[1]]
        
        self.assertEqual(len(severities_exported), 2,
            "Should only export 2 issues (Critical and High)")
        self.assertNotIn('Low', severities_exported,
            "Low severity should be filtered out")


class TestAnalyticsPolish(unittest.TestCase):
    """
    Tests for v3.0.33 Chunk E: Analytics Polish.
    
    Validates:
    - Score trend data retrieval
    - Heatmap data structure
    - API endpoint availability
    """
    
    def test_score_trend_method_exists(self):
        """
        Test that get_score_trend method exists in ScanHistoryDB.
        
        Functionality: Score trends show improvement over time.
        Expects: ScanHistoryDB class has get_score_trend method.
        """
        from scan_history import ScanHistoryDB
        
        # Check method exists
        self.assertTrue(hasattr(ScanHistoryDB, 'get_score_trend'),
            "ScanHistoryDB should have get_score_trend method")
    
    def test_score_trend_returns_correct_format(self):
        """
        Test that score trend returns data in expected format.
        
        API Design: Trend data must be list for charting.
        Expects: get_score_trend returns a list (empty for nonexistent file).
        """
        from scan_history import ScanHistoryDB
        import tempfile
        import os
        
        # Create temp DB for testing
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as f:
            temp_db = f.name
        
        try:
            db = ScanHistoryDB(temp_db)
            
            # Get trend for non-existent file (should return empty list)
            trend = db.get_score_trend('nonexistent.docx', limit=10)
            
            self.assertIsInstance(trend, list,
                "Score trend should return a list")
        finally:
            os.unlink(temp_db)
    
    def test_heatmap_data_structure(self):
        """
        Test that issue data can generate heatmap matrix.
        
        Functionality: Heatmap shows Category × Severity distribution.
        Expects: Matrix correctly counts issues per category-severity pair.
        """
        # Mock issues
        issues = [
            {'severity': 'Critical', 'category': 'Grammar'},
            {'severity': 'High', 'category': 'Grammar'},
            {'severity': 'Medium', 'category': 'Style'},
            {'severity': 'Low', 'category': 'Style'},
            {'severity': 'Info', 'category': 'Formatting'},
        ]
        
        # Build matrix (mimicking JS renderIssueHeatmap logic)
        severities = ['Critical', 'High', 'Medium', 'Low', 'Info']
        matrix = {}
        
        for issue in issues:
            cat = issue.get('category', 'Other')
            sev = issue.get('severity', 'Info')
            
            if cat not in matrix:
                matrix[cat] = {}
            matrix[cat][sev] = matrix[cat].get(sev, 0) + 1
        
        # Verify matrix structure
        self.assertEqual(matrix['Grammar']['Critical'], 1)
        self.assertEqual(matrix['Grammar']['High'], 1)
        self.assertEqual(matrix['Style']['Medium'], 1)
        self.assertEqual(matrix['Style']['Low'], 1)
        self.assertEqual(matrix['Formatting']['Info'], 1)


class TestJobBasedReview(unittest.TestCase):
    """
    Tests for v3.0.39 Batch I: Job-Based Review with Progress.
    
    Validates:
    - /api/review/start endpoint exists
    - Job status endpoint returns progress info
    - Core.py accepts progress callbacks
    """
    
    def setUp(self):
        """Set up test fixtures."""
        try:
            from app import app
            self.app = app
            self.client = app.test_client()
        except ImportError:
            self.skipTest("app module not available")
    
    def test_job_status_endpoint_exists(self):
        """
        Test /api/job/status endpoint exists and reports availability.
        
        Functionality: Frontend needs to know if job system is available.
        Expects: 200 response with available=True.
        """
        response = self.client.get('/api/job/status')
        self.assertEqual(response.status_code, 200)
        
        data = response.get_json()
        self.assertIn('available', data)
        # Job manager should be available
        self.assertTrue(data['available'])
    
    def test_review_start_endpoint_exists(self):
        """
        Test /api/review/start endpoint exists.
        
        Functionality: Job-based review uses dedicated start endpoint.
        Expects: Endpoint exists (400 for missing file, not 404).
        """
        # Get CSRF token first
        csrf_response = self.client.get('/')
        
        # Try to start review without a file (should fail with validation error)
        response = self.client.post(
            '/api/review/start',
            json={'options': {}},
            headers={'X-CSRF-Token': 'test'}
        )
        
        # Should be 400 (validation error) because no file uploaded, not 404
        self.assertIn(response.status_code, [400, 403])
    
    def test_core_accepts_progress_callback(self):
        """
        Test that core.py review_document accepts progress_callback.
        
        Functionality: Progress reporting enables UI progress bars.
        Expects: review_document has progress_callback and cancellation_check params.
        """
        from core import TechWriterReviewEngine
        import inspect
        
        # Get the signature of review_document
        sig = inspect.signature(TechWriterReviewEngine.review_document)
        params = list(sig.parameters.keys())
        
        # Should have progress_callback parameter
        self.assertIn('progress_callback', params,
            "review_document should accept progress_callback parameter")
        
        # Should have cancellation_check parameter
        self.assertIn('cancellation_check', params,
            "review_document should accept cancellation_check parameter")
    
    def test_job_phases_defined(self):
        """
        Test that JobPhase enum has all expected phases.
        
        Functionality: Job phases enable detailed progress reporting.
        Expects: All expected phase names exist in JobPhase enum.
        """
        from job_manager import JobPhase
        
        expected_phases = ['QUEUED', 'EXTRACTING', 'PARSING', 'CHECKING', 
                          'POSTPROCESSING', 'COMPLETE', 'FAILED', 'CANCELLED']
        
        actual_phases = [p.name for p in JobPhase]
        
        for phase in expected_phases:
            self.assertIn(phase, actual_phases,
                f"JobPhase should include {phase}")
    
    def test_job_manager_tracks_progress(self):
        """
        Test that JobManager can track checker progress.
        
        Functionality: Job manager maintains progress state.
        Expects: Job has progress info after update_checker_progress call.
        """
        from job_manager import get_job_manager, JobPhase
        
        manager = get_job_manager()
        
        # Create a test job
        job_id = manager.create_job('test_review', {'filename': 'test.docx'})
        
        # Start job
        manager.start_job(job_id)
        
        # Update to checking phase
        manager.update_phase(job_id, JobPhase.CHECKING, "Running checkers")
        
        # Update checker progress
        manager.update_checker_progress(job_id, 'grammar', 1, 10)
        
        # Verify progress is tracked
        job = manager.get_job(job_id)
        self.assertEqual(job.progress.current_checker, 'grammar')
        self.assertEqual(job.progress.checkers_completed, 1)
        self.assertEqual(job.progress.checkers_total, 10)
        
        # Clean up
        manager.complete_job(job_id, {'issues': []})


class TestStatementToRoleMapping(unittest.TestCase):
    """
    Tests for Statement Forge → Role Responsibilities mapping.
    
    v3.0.41: Batch H - Verifies the bidirectional mapping between
    extracted statements and detected roles.
    """
    
    def setUp(self):
        """Set up test fixtures."""
        try:
            from role_integration import RoleIntegration
            self.integration = RoleIntegration()
        except ImportError:
            self.skipTest("role_integration module not available")
    
    def test_mapping_basic(self):
        """
        Test basic statement-to-role mapping.
        
        Functionality: Statements map to roles by explicit role or text match.
        Expects: Explicit role fields mapped, text matches found, stats correct.
        """
        # Sample statements (as returned by Statement Forge)
        statements = [
            {
                'id': 'stmt1',
                'number': '4.1',
                'description': 'The Systems Engineer shall review all design documents.',
                'role': 'Systems Engineer',
                'directive': 'shall'
            },
            {
                'id': 'stmt2',
                'number': '4.2',
                'description': 'The Quality Manager approves all test procedures.',
                'role': '',
                'directive': ''
            },
            {
                'id': 'stmt3',
                'number': '4.3',
                'description': 'All personnel must complete training.',
                'role': '',
                'directive': 'must'
            }
        ]
        
        # Sample extracted roles
        extracted_roles = {
            'roles': {
                'Systems Engineer': {
                    'variants': ['SE', 'Sys Eng'],
                    'responsibilities': ['Review design documents']
                },
                'Quality Manager': {
                    'variants': ['QM', 'Quality Mgr'],
                    'responsibilities': ['Approve test procedures']
                }
            }
        }
        
        result = self.integration.map_statements_to_roles(statements, extracted_roles)
        
        # Verify structure
        self.assertIn('role_to_statements', result)
        self.assertIn('statement_to_roles', result)
        self.assertIn('unmapped_statements', result)
        self.assertIn('stats', result)
        
        # Verify mappings
        self.assertIn('Systems Engineer', result['role_to_statements'])
        self.assertIn('Quality Manager', result['role_to_statements'])
        
        # stmt1 should map to Systems Engineer (explicit role)
        se_stmts = result['role_to_statements']['Systems Engineer']
        self.assertEqual(len(se_stmts), 1)
        self.assertEqual(se_stmts[0]['id'], 'stmt1')
        
        # stmt2 should map to Quality Manager (found in description)
        qm_stmts = result['role_to_statements']['Quality Manager']
        self.assertEqual(len(qm_stmts), 1)
        self.assertEqual(qm_stmts[0]['id'], 'stmt2')
        
        # stmt3 should be unmapped (no role match)
        self.assertIn('stmt3', result['unmapped_statements'])
        
        # Verify stats
        self.assertEqual(result['stats']['total_statements'], 3)
        self.assertEqual(result['stats']['mapped_statements'], 2)
    
    def test_mapping_empty_roles(self):
        """
        Test mapping with no roles available.
        
        Edge Case: No roles means all statements unmapped.
        Expects: All statements in unmapped list, mapped_statements=0.
        """
        statements = [{'id': 'stmt1', 'description': 'Test statement', 'role': ''}]
        
        result = self.integration.map_statements_to_roles(statements, {'roles': {}})
        
        self.assertEqual(result['stats']['total_statements'], 1)
        self.assertEqual(result['stats']['mapped_statements'], 0)
        self.assertIn('stmt1', result['unmapped_statements'])
    
    def test_mapping_empty_statements(self):
        """
        Test mapping with no statements.
        
        Edge Case: No statements produces empty mapping with 0% coverage.
        Expects: total_statements=0, coverage_percent=0.
        """
        extracted_roles = {
            'roles': {
                'Test Role': {'variants': [], 'responsibilities': []}
            }
        }
        
        result = self.integration.map_statements_to_roles([], extracted_roles)
        
        self.assertEqual(result['stats']['total_statements'], 0)
        self.assertEqual(result['stats']['coverage_percent'], 0)
    
    def test_mapping_variant_matching(self):
        """
        Test that role variants are matched.
        
        Functionality: Role variants like 'Sys Eng' should match 'Systems Engineer'.
        Expects: Statement with variant matched to canonical role name.
        """
        statements = [
            {
                'id': 'stmt1',
                'description': 'The Sys Eng coordinates with the design team.',
                'role': ''
            }
        ]
        
        extracted_roles = {
            'roles': {
                'Systems Engineer': {
                    'variants': ['Sys Eng', 'System Engineer'],
                    'responsibilities': []
                }
            }
        }
        
        result = self.integration.map_statements_to_roles(statements, extracted_roles)
        
        # Should match via variant 'Sys Eng'
        se_stmts = result['role_to_statements']['Systems Engineer']
        self.assertEqual(len(se_stmts), 1)
        self.assertNotIn('stmt1', result['unmapped_statements'])
    
    def test_mapping_api_endpoint_exists(self):
        """
        Test that the mapping API endpoint is registered.
        
        Functionality: Frontend can request statement-to-role mapping.
        Expects: POST /api/statement-forge/map-to-roles returns JSON (not 404).
        """
        try:
            from app import app
            client = app.test_client()
            
            # Endpoint should exist (even if it returns error without session)
            response = client.post('/api/statement-forge/map-to-roles')
            
            # Should get JSON response (not 404)
            self.assertNotEqual(response.status_code, 404)
            data = response.get_json()
            self.assertIsNotNone(data)
            
        except ImportError:
            self.skipTest("app module not available")
    
    def test_mapping_status_endpoint_exists(self):
        """
        Test that the mapping status endpoint is registered.
        
        Functionality: Frontend checks if mapping is available.
        Expects: GET endpoint returns availability status fields.
        """
        try:
            from app import app
            client = app.test_client()
            
            response = client.get('/api/statement-forge/role-mapping-status')
            
            self.assertNotEqual(response.status_code, 404)
            data = response.get_json()
            self.assertIn('statements_available', data)
            self.assertIn('roles_available', data)
            self.assertIn('can_map', data)
            
        except ImportError:
            self.skipTest("app module not available")


class TestUIPolish(unittest.TestCase):
    """
    v3.0.43: Tests for UI Polish features - Essentials Mode, Run-State Indicator.
    """
    
    def test_essentials_mode_css_exists(self):
        """
        Test that essentials-mode CSS rules exist.
        
        Functionality: Essentials mode hides advanced features.
        Expects: CSS files contain .essentials-mode and related selectors.
        """
        css_dir = Path(__file__).parent / 'static' / 'css'
        if not css_dir.exists():
            self.skipTest("CSS directory not available")
        
        # Read all CSS files
        css_content = ""
        for css_file in css_dir.glob('*.css'):
            css_content += css_file.read_text()
        
        self.assertIn('.essentials-mode', css_content)
        self.assertIn('#super-tools', css_content)
        self.assertIn('#btn-statement-forge', css_content)
        self.assertIn('#btn-triage-mode', css_content)
    
    def test_run_state_indicator_css_exists(self):
        """
        Test that run-state indicator CSS rules exist.
        
        Functionality: Run-state indicator shows analysis progress.
        Expects: CSS files contain .run-state-indicator and animation.
        """
        css_dir = Path(__file__).parent / 'static' / 'css'
        if not css_dir.exists():
            self.skipTest("CSS directory not available")
        
        # Read all CSS files
        css_content = ""
        for css_file in css_dir.glob('*.css'):
            css_content += css_file.read_text()
        
        self.assertIn('.run-state-indicator', css_content)
        self.assertIn('.run-state-pulse', css_content)
        self.assertIn('@keyframes run-state-pulse', css_content)
    
    def test_essentials_mode_toggle_in_html(self):
        """
        Test that essentials mode toggle exists in settings.
        
        UI: Settings panel must have essentials mode toggle.
        Expects: HTML contains settings-essentials-mode ID and label text.
        """
        html_path = Path(__file__).parent / 'templates' / 'index.html'
        if not html_path.exists():
            self.skipTest("HTML template not available")
        
        html_content = html_path.read_text()
        self.assertIn('settings-essentials-mode', html_content)
        self.assertIn('Essentials mode', html_content)
    
    def test_run_state_indicator_in_html(self):
        """
        Test that run-state indicator exists in header.
        
        UI: Header must have run-state indicator elements.
        Expects: HTML contains run-state-indicator, run-state-text, run-state-cancel.
        """
        html_path = Path(__file__).parent / 'templates' / 'index.html'
        if not html_path.exists():
            self.skipTest("HTML template not available")
        
        html_content = html_path.read_text()
        self.assertIn('run-state-indicator', html_content)
        self.assertIn('run-state-text', html_content)
        self.assertIn('run-state-cancel', html_content)
    
    def test_essentials_mode_in_state_settings(self):
        """
        Test that essentialsMode is in default State settings.
        
        JavaScript: State must include essentialsMode default.
        Expects: app.js contains essentialsMode: false.
        """
        js_path = Path(__file__).parent / 'static' / 'js' / 'app.js'
        if not js_path.exists():
            self.skipTest("JS file not available")
        
        js_content = js_path.read_text()
        self.assertIn('essentialsMode:', js_content)
        # Should default to false
        self.assertIn('essentialsMode: false', js_content)
    
    def test_search_functionality_exists(self):
        """
        Test that issue search input and logic exist.
        
        UI: Issue search enables finding specific problems.
        Expects: HTML has search input, JS has search handler.
        """
        html_path = Path(__file__).parent / 'templates' / 'index.html'
        if not html_path.exists():
            self.skipTest("HTML template not available")
        
        html_content = html_path.read_text()
        self.assertIn('issue-search', html_content)
        self.assertIn('btn-clear-search', html_content)
        
        js_path = Path(__file__).parent / 'static' / 'js' / 'app.js'
        js_content = js_path.read_text()
        self.assertIn("getElementById('issue-search')", js_content)


class TestFixAssistantV2API(unittest.TestCase):
    """
    Test Fix Assistant v2 API endpoints.
    
    v3.0.103: Validates FAV2 endpoints for pattern learning, prediction,
    statistics, dictionary management, and report generation.
    """
    
    def setUp(self):
        """Set up test client with CSRF token."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
        # Get CSRF token for POST requests
        response = self.client.get('/api/csrf-token')
        self.csrf_token = json.loads(response.data).get('csrf_token')
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
    
    def test_learner_record_endpoint_exists(self):
        """
        Test /api/learner/record endpoint accepts POST requests.
        
        Fix Assistant v2: Records user decisions for pattern learning.
        Expects: Endpoint exists and accepts valid decision data (not 404).
        """
        response = self.client.post(
            '/api/learner/record',
            headers={'X-CSRF-Token': self.csrf_token},
            json={
                'issue_id': 'test-001',
                'decision': 'accept',
                'category': 'grammar',
                'pattern': 'test pattern'
            },
            content_type='application/json'
        )
        # Should not be 404 (endpoint exists)
        self.assertNotEqual(response.status_code, 404,
            "/api/learner/record endpoint should exist")
    
    def test_learner_predict_endpoint_exists(self):
        """
        Test /api/learner/predict endpoint accepts POST requests.
        
        Fix Assistant v2: Predicts user decision based on past patterns.
        Expects: Endpoint exists and returns prediction data (not 404).
        """
        response = self.client.post(
            '/api/learner/predict',
            headers={'X-CSRF-Token': self.csrf_token},
            json={
                'category': 'grammar',
                'pattern': 'test pattern'
            },
            content_type='application/json'
        )
        self.assertNotEqual(response.status_code, 404,
            "/api/learner/predict endpoint should exist")
    
    def test_learner_stats_endpoint_exists(self):
        """
        Test /api/learner/statistics endpoint returns statistics.
        
        Fix Assistant v2: Returns learning statistics for UI display.
        Expects: 200 response with stats data structure.
        """
        response = self.client.get('/api/learner/statistics')
        self.assertEqual(response.status_code, 200,
            "/api/learner/statistics endpoint should return 200")
        data = json.loads(response.data)
        self.assertIn('success', data)
    
    def test_learner_dictionary_endpoint_exists(self):
        """
        Test /api/learner/dictionary endpoint for custom terms.
        
        Fix Assistant v2: Manages user's custom skip dictionary.
        Expects: 200 response with dictionary data.
        """
        response = self.client.get('/api/learner/dictionary')
        self.assertEqual(response.status_code, 200,
            "/api/learner/dictionary endpoint should return 200")
    
    def test_report_generate_endpoint_exists(self):
        """
        Test /api/report/generate endpoint accepts POST requests.
        
        Fix Assistant v2: Generates PDF summary reports.
        Expects: Endpoint exists (may require session data to succeed, not 404).
        """
        response = self.client.post(
            '/api/report/generate',
            headers={'X-CSRF-Token': self.csrf_token},
            json={'format': 'pdf'},
            content_type='application/json'
        )
        # Should not be 404 (endpoint exists)
        self.assertNotEqual(response.status_code, 404,
            "/api/report/generate endpoint should exist")


class TestBatchLimits(unittest.TestCase):
    """
    Test batch upload limits (v3.0.102 security feature).
    
    v3.0.103: Validates batch upload endpoint and limit constants.
    """
    
    def setUp(self):
        """Set up test client."""
        app.config['TESTING'] = True
        self.client = app.test_client()
        self.ctx = app.app_context()
        self.ctx.push()
        response = self.client.get('/api/csrf-token')
        self.csrf_token = json.loads(response.data).get('csrf_token')
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
    
    def test_batch_endpoint_exists(self):
        """
        Test /api/upload/batch endpoint exists.
        
        v3.0.102: Batch upload with file count and size limits.
        Expects: Endpoint exists and validates input (400 for no files, not 404).
        """
        response = self.client.post(
            '/api/upload/batch',
            headers={'X-CSRF-Token': self.csrf_token},
            content_type='multipart/form-data'
        )
        # 400 (no files) is expected, 404 would mean endpoint missing
        self.assertNotEqual(response.status_code, 404,
            "/api/upload/batch endpoint should exist")
    
    def test_batch_constants_defined(self):
        """
        Test batch limit constants are defined in app.
        
        v3.0.102: MAX_BATCH_SIZE=10, MAX_BATCH_TOTAL_SIZE=100MB.
        Expects: Constants exist with reasonable values (or skip if not yet implemented).
        """
        try:
            from app import MAX_BATCH_SIZE, MAX_BATCH_TOTAL_SIZE
            self.assertEqual(MAX_BATCH_SIZE, 10,
                "MAX_BATCH_SIZE should be 10")
            self.assertEqual(MAX_BATCH_TOTAL_SIZE, 100 * 1024 * 1024,
                "MAX_BATCH_TOTAL_SIZE should be 100MB")
        except ImportError:
            # Constants may not exist yet - this is acceptable
            self.skipTest("Batch limit constants not yet defined in app.py")


class TestSessionCleanup(unittest.TestCase):
    """
    Test automatic session cleanup (v3.0.102 feature).
    
    v3.0.103: Validates SessionManager class and cleanup capabilities.
    """
    
    def setUp(self):
        """Set up test environment."""
        app.config['TESTING'] = True
        self.ctx = app.app_context()
        self.ctx.push()
    
    def tearDown(self):
        """Clean up."""
        self.ctx.pop()
    
    def test_session_manager_exists(self):
        """
        Test SessionManager class is available.
        
        v3.0.102: SessionManager handles session lifecycle.
        Expects: SessionManager class exists and can be imported.
        """
        from app import SessionManager
        self.assertTrue(callable(SessionManager),
            "SessionManager should be a callable class")
    
    def test_session_manager_has_cleanup(self):
        """
        Test SessionManager has cleanup capability.
        
        v3.0.102: Automatic cleanup removes stale sessions.
        Expects: cleanup_old or cleanup_old_sessions method exists.
        """
        from app import SessionManager
        # Check for either cleanup_old or cleanup_old_sessions
        has_cleanup = (hasattr(SessionManager, 'cleanup_old_sessions') or 
                       hasattr(SessionManager, 'cleanup_old'))
        self.assertTrue(has_cleanup,
            "SessionManager should have cleanup_old or cleanup_old_sessions method")


def run_tests():
    """Run all tests and return results."""
    loader = unittest.TestLoader()
    suite = unittest.TestSuite()
    
    # Add test classes
    test_classes = [
        TestSecurityControls,
        TestFileValidation,
        TestAPIEndpoints,
        TestErrorHandling,
        TestSessionManagement,
        TestVersionConsistency,
        TestCodeQuality,
        TestConfigDefaults,
        TestRateLimitIntegration,
        TestAuthenticationIntegration,
        TestAPIKeyAuth,
        TestRolesGraphAPI,
        TestCapabilitiesAPI,
        TestStaticFileSecurity,
        TestAcronymChecker,
        TestRoleDeliverableSeparation,  # v3.0.12: Role/deliverable separation tests
        TestHyperlinkHealth,  # v3.0.33 Chunk B: Hyperlink health tests
        TestHyperlinkConfigEndpoint,  # v3.0.37: Hyperlink config endpoint tests
        TestCommentInserter,  # v3.0.37 Batch G: Comment inserter tests
        TestStatementForgePersistence,  # v3.0.33 Chunk C: SF persistence tests
        TestExportEnhancements,  # v3.0.33 Chunk D: Export enhancement tests
        TestAnalyticsPolish,  # v3.0.33 Chunk E: Analytics polish tests
        TestJobBasedReview,  # v3.0.39 Batch I: Job-based review tests
        TestStatementToRoleMapping,  # v3.0.41 Batch H: SF→Role mapping tests
        TestUIPolish,  # v3.0.43: UI Polish tests (Essentials Mode, Run-State)
        TestFixAssistantV2API,  # v3.0.103: Fix Assistant v2 API tests
        TestBatchLimits,  # v3.0.103: Batch upload limit tests
        TestSessionCleanup,  # v3.0.103: Session cleanup tests
    ]
    
    for test_class in test_classes:
        suite.addTests(loader.loadTestsFromTestCase(test_class))
    
    # Run tests with verbosity
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    return result


def run_guardrail_checks():
    """Run static guardrail checks that should fail CI."""
    print("\n" + "="*60)
    print("GUARDRAIL CHECKS")
    print("="*60 + "\n")
    
    failures = []
    base_path = Path(__file__).parent
    
    # Check 1: No bare except
    print("1. Checking for bare 'except:' statements...")
    import re
    for py_file in base_path.glob('*.py'):
        with open(py_file, 'r') as f:
            content = f.read()
        matches = re.findall(r'^\s*except:\s*$', content, re.MULTILINE)
        if matches:
            failures.append(f"   FAIL: {py_file.name} has {len(matches)} bare except clauses")
    if not any('bare except' in f for f in failures):
        print("   PASS: No bare except clauses found")
    
    # Check 2: No literal debug=True anywhere (STRICT GUARDRAIL)
    print("\n2. Checking for literal 'debug=True'...")
    debug_failures = []
    for py_file in base_path.glob('*.py'):
        # Skip test file itself for this check
        if py_file.name == 'tests.py':
            continue
        with open(py_file, 'r') as f:
            lines = f.readlines()
        for i, line in enumerate(lines, 1):
            # Check for literal debug=True (not debug=variable)
            if re.search(r'\bdebug\s*=\s*True\b', line):
                # This is a FAILURE - no literal debug=True allowed
                debug_failures.append(f"   FAIL: {py_file.name}:{i} contains literal 'debug=True'")
    
    if debug_failures:
        failures.extend(debug_failures)
        for f in debug_failures:
            print(f)
    else:
        print("   PASS: No literal 'debug=True' found")
    
    # Check 3: Version consistency
    print("\n3. Checking version consistency...")
    versions = set()
    for py_file in base_path.glob('*.py'):
        with open(py_file, 'r') as f:
            content = f.read()
        matches = re.findall(r'^__version__\s*=\s*["\']([^"\']+)["\']', content, re.MULTILINE)
        versions.update(matches)
    
    if len(versions) == 1:
        print(f"   PASS: Single version found: {versions.pop()}")
    elif len(versions) > 1:
        failures.append(f"   FAIL: Multiple versions found: {versions}")
        print(f"   FAIL: Multiple versions found: {versions}")
    else:
        print("   INFO: No __version__ strings found in root modules")
    
    # Check 4: Required files exist
    print("\n4. Checking required files...")
    required_files = ['app.py', 'config_logging.py', 'core.py', 'index.html', 'app.js', 'style.css']
    for req_file in required_files:
        if (base_path / req_file).exists():
            print(f"   PASS: {req_file} exists")
        else:
            failures.append(f"   FAIL: {req_file} missing")
            print(f"   FAIL: {req_file} missing")
    
    # Check 5: Rate limiting is enabled by default
    print("\n5. Checking rate limiting default...")
    cfg = get_config()
    if cfg.rate_limit_enabled:
        print("   PASS: Rate limiting enabled by default")
    else:
        failures.append("   FAIL: Rate limiting not enabled by default")
        print("   FAIL: Rate limiting not enabled by default")
    
    # Check 6: Auth defaults to disabled (safe for local dev)
    print("\n6. Checking auth default...")
    if not cfg.auth_enabled:
        print("   PASS: Auth disabled by default (safe for local dev)")
    else:
        print("   INFO: Auth enabled by default")
    
    print("\n" + "="*60)
    if failures:
        print(f"GUARDRAIL CHECKS: {len(failures)} FAILURES")
        for f in failures:
            print(f)
        return False
    else:
        print("GUARDRAIL CHECKS: ALL PASSED")
        return True


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser(description='TechWriterReview Test Suite')
    parser.add_argument('--guardrails', action='store_true', help='Run guardrail checks only')
    parser.add_argument('--all', action='store_true', help='Run all tests and guardrails')
    args = parser.parse_args()
    
    if args.guardrails:
        success = run_guardrail_checks()
        sys.exit(0 if success else 1)
    elif args.all:
        result = run_tests()
        guardrails_ok = run_guardrail_checks()
        sys.exit(0 if result.wasSuccessful() and guardrails_ok else 1)
    else:
        result = run_tests()
        sys.exit(0 if result.wasSuccessful() else 1)
